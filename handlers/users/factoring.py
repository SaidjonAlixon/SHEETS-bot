from aiogram import types, F
from aiogram.types import FSInputFile, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from loader import dp, bot
from keyboards.default.factoring_menu import factoring_menu
from keyboards.default.main_menu import get_main_menu, load_select_menu
from services.excel_parser import ExcelParser
from services.google_sheets import get_sheet_service
from states.bot_states import BotStates
from utils.company_storage import get_company
import pandas as pd
import os
import re


@dp.message(F.text == "📄 Factoring Payments")
async def enter_factoring(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=load_select_menu)
        return
    await state.set_state(BotStates.Factoring)
    await message.answer("Factoring Payments bo'limiga xush kelibsiz.\n"
                         "Iltimos, operatsiyani tanlang:", reply_markup=factoring_menu)

@dp.message(F.text == "⬅️ Orqaga", BotStates.Factoring)
async def go_back(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Asosiy menyu:", reply_markup=get_main_menu(message.from_user.id))

@dp.message(F.text == "⬅️ Orqaga", BotStates.FactoringDateRange)
async def go_back_from_date(message: types.Message, state: FSMContext):
    await state.set_state(BotStates.Factoring)
    await state.clear()
    await message.answer("Factoring Payments:", reply_markup=factoring_menu)

@dp.message(F.text == "⬅️ Orqaga", BotStates.FactoringSearchLoad)
async def go_back_from_search(message: types.Message, state: FSMContext):
    await state.clear()
    await state.set_state(BotStates.Factoring)
    await message.answer("Factoring Payments:", reply_markup=factoring_menu)

# ——— 📋 Oxirgi yuklamalar ———
@dp.message(F.text == "📋 Oxirgi yuklamalar", BotStates.Factoring)
async def factoring_recent_loads(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=load_select_menu)
        return
    try:
        sheet_service = get_sheet_service()
        all_sheets = sheet_service.get_all_sheet_names(company)
    except Exception as e:
        if "429" in str(e):
            await message.answer("⚠️ Google Sheets limiti tugadi. 1–2 daqiqa kutib qayta urinib ko'ring.")
        else:
            await message.answer(f"Sheetga ulanish xatosi: {e}")
        return
    if not all_sheets:
        await message.answer("Listlar topilmadi.")
        return
    buttons = []
    row = []
    for i, name in enumerate(all_sheets):
        row.append(InlineKeyboardButton(text=name, callback_data=f"recent:{i}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    await message.answer(
        "<b>Oxirgi yuklamalar</b>\nQaysi listdan ko'rsatay?", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

@dp.callback_query(F.data.startswith("recent:"))
async def callback_recent_sheet(callback: types.CallbackQuery):
    await callback.answer()
    company = get_company(callback.from_user.id)
    if not company:
        await callback.message.edit_text("Iltimos, /start bosing va Load tanlang.")
        return
    try:
        idx = int(callback.data.replace("recent:", "").strip())
    except ValueError:
        return
    await callback.message.edit_text("⏳ Kutib turing, natija tez orada chiqadi...")
    try:
        sheet_service = get_sheet_service()
        all_sheets = sheet_service.get_all_sheet_names(company)
    except Exception as e:
        await callback.message.edit_text(f"Xatolik: {e}")
        return
    if idx < 0 or idx >= len(all_sheets):
        return
    sheet_name = all_sheets[idx]
    loads = [x for x in sheet_service.get_recent_loads(sheet_name, limit=15, company=company) if x]
    if not loads:
        await callback.message.edit_text(f"List <b>{sheet_name}</b> da yuklar topilmadi.")
        return
    lines = [f"📋 <b>Oxirgi yuklamalar ({sheet_name})</b>\n"]
    for i, r in enumerate(reversed(loads), 1):
        if not r:
            continue
        ln = r.get('load_number') or '-'
        dr = (r.get('driver') or '-')[:20]
        inv = r.get('invoiced') or '-'
        paid = r.get('broker_paid') or '-'
        st = r.get('status') or '-'
        lines.append(f"{i}. LOAD #{ln} | {dr} | Inv: {inv} | Paid: {paid} | {st}")
    text = "\n".join(lines)
    if len(text) > 4000:
        text = "\n".join(lines[:25]) + "\n\n... (faqat 25 ta ko'rsatildi)"
    await callback.message.edit_text(text)

# ——— 🔍 Yuk raqami bo'yicha qidirish ———
@dp.message(F.text == "🔍 Yuk raqami bo'yicha qidirish", BotStates.Factoring)
async def factoring_search_ask(message: types.Message, state: FSMContext):
    await state.set_state(BotStates.FactoringSearchLoad)
    await message.answer("Yuk raqamini kiriting (masalan: 30158135):")

@dp.message(F.text, BotStates.FactoringSearchLoad)
async def factoring_search_run(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=load_select_menu)
        return
    load_num = message.text.strip()
    if not load_num:
        await message.answer("Yuk raqamini kiriting.")
        return
    try:
        sheet_service = get_sheet_service()
        row_num, sheet_name = sheet_service.find_load_sync(load_num, company=company)
    except Exception as e:
        await message.answer(f"Xatolik: {e}")
        await state.set_state(BotStates.Factoring)
        return
    if not row_num or not sheet_name:
        await message.answer(f"❌ Yuk raqami <b>{load_num}</b> hech qanday listda topilmadi.", reply_markup=factoring_menu)
        await state.set_state(BotStates.Factoring)
        return
    row = sheet_service.get_row_display(row_num, sheet_name, company)
    await state.set_state(BotStates.Factoring)
    if not row:
        await message.answer("Ma'lumot o'qish xatosi.", reply_markup=factoring_menu)
        return
    text = (
        f"✅ <b>Topildi</b> (list: {sheet_name})\n\n"
        f"LOAD #: {row.get('load_number') or '-'}\n"
        f"Haydovchi: {row.get('driver') or '-'}\n"
        f"PU Date: {row.get('pu_date') or '-'}\n"
        f"INVOICED AMOUNT: {row.get('invoiced') or '-'}\n"
        f"BROKER PAID: {row.get('broker_paid') or '-'}\n"
        f"STATUS: {row.get('status') or '-'}"
    )
    await message.answer(text, reply_markup=factoring_menu)

# ——— 📊 Hisobot olish ———
@dp.message(F.text == "📊 Hisobot olish", BotStates.Factoring)
async def factoring_report_ask(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=load_select_menu)
        return
    try:
        sheet_service = get_sheet_service()
        all_sheets = sheet_service.get_all_sheet_names(company)
    except Exception as e:
        if "429" in str(e):
            await message.answer("⚠️ Google Sheets limiti tugadi. 1–2 daqiqa kutib qayta urinib ko'ring.")
        else:
            await message.answer(f"Sheetga ulanish xatosi: {e}")
        return
    if not all_sheets:
        await message.answer("Listlar topilmadi.")
        return
    buttons = []
    row = []
    for i, name in enumerate(all_sheets):
        row.append(InlineKeyboardButton(text=name, callback_data=f"report:{i}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    await message.answer(
        "<b>Hisobot olish</b>\nQaysi list bo'yicha hisobot?", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

@dp.callback_query(F.data.startswith("report:"))
async def callback_report_sheet(callback: types.CallbackQuery):
    await callback.answer()
    company = get_company(callback.from_user.id)
    if not company:
        await callback.message.edit_text("Iltimos, /start bosing va Load tanlang.")
        return
    try:
        idx = int(callback.data.replace("report:", "").strip())
    except ValueError:
        return
    await callback.message.edit_text("⏳ Kutib turing, natija tez orada chiqadi...")
    try:
        sheet_service = get_sheet_service()
        all_sheets = sheet_service.get_all_sheet_names(company)
    except Exception as e:
        await callback.message.edit_text(f"Xatolik: {e}")
        return
    if idx < 0 or idx >= len(all_sheets):
        return
    sheet_name = all_sheets[idx]
    summary = sheet_service.get_sheet_summary(sheet_name, company)
    if not summary:
        await callback.message.edit_text(f"List <b>{sheet_name}</b> bo'yicha hisobot olish xatosi.")
        return
    text = (
        f"📊 <b>Hisobot: {sheet_name}</b>\n\n"
        f"Yuklar soni: {summary['count']}\n"
        f"Jami INVOICED AMOUNT: ${summary['total_invoiced']:,.2f}\n"
        f"Jami BROKER PAID: ${summary['total_broker_paid']:,.2f}"
    )
    await callback.message.edit_text(text)

@dp.message(F.text == "📤 Fayl yuklash (Excel xlsx/xls)", BotStates.Factoring)
async def ask_file(message: types.Message):
    await message.answer("Iltimos, Factoring Invoice faylini (Excel xlsx, xls) yuboring.\n"
                         "Faylda Load/PO # va Invoice Amount ustunlari bo'lishi kerak.")

@dp.message(F.document, BotStates.Factoring)
async def handle_factoring_document(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=load_select_menu)
        return
    document = message.document
    file_id = document.file_id
    file_name = document.file_name

    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        await message.answer("Iltimos, faqat Excel (xlsx, xls) fayl yuklang.")
        return

    await message.answer("Fayl qabul qilindi. Yuklanmoqda... ⏳")

    file = await bot.get_file(file_id)
    file_content = await bot.download_file(file.file_path)
    content_bytes = file_content.read()

    parsed_data = ExcelParser.parse_factoring_report(content_bytes)
    if not parsed_data:
        parsed_data = ExcelParser.parse_invoice(content_bytes)

    if not parsed_data:
        await message.answer("Fayldan ma'lumot o'qib bo'lmadi. D (Load/PO #) va E (Invoice Amount) ustunlarini tekshiring.")
        return

    all_sheets = []
    try:
        sheet_service = get_sheet_service()
        all_sheets = sheet_service.get_all_sheet_names(company)
    except Exception:
        pass

    await state.update_data(
        factoring_file=content_bytes,
        factoring_filename=file_name,
        factoring_parsed=parsed_data,
        factoring_sheet_names=all_sheets,
        selected_company=company
    )
    await state.set_state(BotStates.FactoringDateRange)

    if all_sheets:
        # Inline tugmalar: har bir list uchun tugma (2 ta qatorda)
        buttons = []
        row = []
        for i, name in enumerate(all_sheets):
            # callback_data 64 belgidan oshmasligi kerak; indeks orqali yuboramiz
            row.append(InlineKeyboardButton(text=name, callback_data=f"fs:{i}"))
            if len(row) == 2:
                buttons.append(row)
                row = []
        if row:
            buttons.append(row)
        kb = InlineKeyboardMarkup(inline_keyboard=buttons)
        await message.answer(
            f"✅ Fayl qabul qilindi ({len(parsed_data)} ta yozuv).\n\n"
            f"<b>Qaysi listni tanlaysiz?</b> Quyidagi tugmalardan birini bosing:",
            reply_markup=kb
        )
    else:
        await message.answer(
            f"✅ Fayl qabul qilindi ({len(parsed_data)} ta yozuv).\n\n"
            f"<b>Qaysi oydan qaysi oygacha tekshiray?</b>\n"
            f"List nomini yozing (masalan: 10.08-10.15)",
            reply_markup=factoring_menu
        )


async def _process_factoring_sheet(sheet_name: str, parsed_data: list, file_name: str, state: FSMContext, chat_id: int, company: str):
    """Tanlangan list uchun factoring ma'lumotlarini sheetga yozish. Batch — Fuel/Toll kabi tez."""
    try:
        sheet_service = get_sheet_service()
    except Exception as e:
        if "429" in str(e):
            await bot.send_message(chat_id, "⚠️ Google Sheets limiti tugadi. 1–2 daqiqa kutib qayta urinib ko'ring.")
        else:
            await bot.send_message(chat_id, f"Xatolik: {e}")
        return

    ws = sheet_service.get_load_board(sheet_name, company)
    if not ws:
        await bot.send_message(chat_id, f"❌ «{sheet_name}» nomli list topilmadi.")
        return

    updated_count, skipped_count, not_found_count, results = sheet_service.update_factoring_batch(
        sheet_name, parsed_data, company=company
    )

    await state.clear()
    await state.set_state(BotStates.Factoring)

    result_df = pd.DataFrame(results)
    import tempfile
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    report_filename = tmp.name
    tmp.close()
    result_df.to_excel(report_filename, index=False)

    # Status ustunini ranglash: UPDATED=yashil, SKIPPED=sariq, LOAD NOT FOUND=qizil
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill

        wb = load_workbook(report_filename)
        ws_rep = wb.active
        status_col = None
        for c in range(1, ws_rep.max_column + 1):
            if str(ws_rep.cell(row=1, column=c).value).strip() == "Status":
                status_col = c
                break
        if status_col:
            green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            yellow_fill = PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid")
            red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
            for r in range(2, ws_rep.max_row + 1):
                cell = ws_rep.cell(row=r, column=status_col)
                v = str(cell.value or "").upper()
                if "UPDATED" in v:
                    cell.fill = green_fill
                elif "SKIPPED" in v:
                    cell.fill = yellow_fill
                elif "NOT FOUND" in v or "EMPTY" in v:
                    cell.fill = red_fill
            wb.save(report_filename)
    except Exception:
        pass

    await bot.send_message(
        chat_id,
        f"✅ Factoring yakunlandi: <b>{sheet_name}</b>\n\n"
        f"Yangilandi: {updated_count}\nO'tkazib yuborildi: {skipped_count}\nTopilmadi: {not_found_count}",
        reply_markup=factoring_menu
    )
    await bot.send_document(chat_id, types.FSInputFile(report_filename))
    os.remove(report_filename)


@dp.callback_query(F.data.startswith("fs:"), BotStates.FactoringDateRange)
async def handle_factoring_sheet_select(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    data = await state.get_data()
    company = data.get("selected_company") or get_company(callback.from_user.id)
    if not company:
        await callback.message.edit_text("Iltimos, /start bosing va Load tanlang.")
        return
    idx_str = callback.data.replace("fs:", "").strip()
    try:
        idx = int(idx_str)
    except ValueError:
        return
    sheet_names = data.get("factoring_sheet_names") or []
    parsed_data = data.get("factoring_parsed")
    file_name = data.get("factoring_filename", "report.xlsx")
    if idx < 0 or idx >= len(sheet_names) or not parsed_data:
        await callback.message.edit_text("❌ Ma'lumotlar eski. Qaytadan fayl yuboring.", reply_markup=None)
        await state.set_state(BotStates.Factoring)
        return
    sheet_name = sheet_names[idx]
    await callback.message.edit_text("⏳ Kutib turing, natija tez orada chiqadi...", reply_markup=None)
    await _process_factoring_sheet(sheet_name, parsed_data, file_name, state, callback.message.chat.id, company)


@dp.message(F.text, BotStates.FactoringDateRange)
async def handle_factoring_date_range(message: types.Message, state: FSMContext):
    sheet_name = message.text.strip()
    if not sheet_name:
        await message.answer("Iltimos, list nomini kiriting yoki yuqoridagi tugmalardan birini bosing.")
        return

    data = await state.get_data()
    parsed_data = data.get('factoring_parsed')
    file_name = data.get('factoring_filename', 'report.xlsx')
    company = data.get("selected_company") or get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=load_select_menu)
        return

    if not parsed_data:
        await state.set_state(BotStates.Factoring)
        await message.answer("Fayl ma'lumotlari topilmadi. Qaytadan fayl yuboring.", reply_markup=factoring_menu)
        return

    await _process_factoring_sheet(sheet_name, parsed_data, file_name, state, message.chat.id, company)
