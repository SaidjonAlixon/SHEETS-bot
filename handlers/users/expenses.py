from aiogram import types, F
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from loader import dp, bot
from keyboards.default.sub_menus import expenses_menu
from keyboards.default.main_menu import get_main_menu, get_load_select_menu
from states.bot_states import BotStates
from utils.company_storage import get_company


def _fuel_norm_header(val) -> str:
    import re
    import pandas as pd
    if val is None or (isinstance(val, float) and val != val):
        return ""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    s = s.replace("\ufeff", "").replace("\xa0", " ")
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _flatten_column_name(col) -> str:
    """MultiIndex yoki tuple ustun nomlarini bitta qatorga."""
    if isinstance(col, tuple):
        parts = [str(p) for p in col if p is not None and str(p) != "nan" and str(p).strip()]
        return " ".join(parts)
    return str(col)


def match_fuel_four_columns(headers: list[str]) -> tuple[int, int, int, int] | None:
    """
    Normalizatsiyalangan sarlavha ro'yxatidan ustun indekslari: card, date, disc, amt.
    """
    n = len(headers)
    if n < 4:
        return None
    card_col = date_col = disc_col = amt_col = None
    for j, h in enumerate(headers):
        if not h:
            continue
        if card_col is None and "card" in h:
            card_col = j
        if date_col is None:
            if ("tran" in h or "trans" in h) and "date" in h:
                date_col = j
            elif h in ("transaction date", "post date") and "toll" not in h:
                date_col = j
        if disc_col is None:
            if (
                "disc" in h
                and "amt" in h
                and "ppu" not in h
                and "cost" not in h
                and "type" not in h
            ):
                disc_col = j
        if amt_col is None:
            if h == "amt" or (h == "amount" and "disc" not in h):
                amt_col = j
            elif (
                h.endswith(" amt")
                and "disc" not in h
                and "ppu" not in h
                and h not in ("disc amt",)
            ):
                amt_col = j
    if date_col is None:
        for j, h in enumerate(headers):
            if h == "date":
                date_col = j
                break
    if amt_col is None:
        for j, h in enumerate(headers):
            if h in ("net amt", "fuel amt", "total amt", "total") and "disc" not in h:
                amt_col = j
                break
    if (
        card_col is not None
        and date_col is not None
        and disc_col is not None
        and amt_col is not None
        and len({card_col, date_col, disc_col, amt_col}) == 4
    ):
        return card_col, date_col, disc_col, amt_col
    return None


def find_fuel_transaction_header_map(df) -> tuple[int, int, int, int, int] | None:
    """
    Excel jadvalida (header=None) ustun nomlari qatorini qidiradi.
    Qaytaradi: (header_row_index, card_col, date_col, disc_col, amt_col) yoki None.
    """
    import pandas as pd
    if df is None or df.empty:
        return None
    ncols = int(df.shape[1])
    if ncols < 4:
        return None
    max_hr = min(45, len(df))
    for hr in range(max_hr):
        headers = [_fuel_norm_header(df.iloc[hr, j]) for j in range(ncols)]
        m = match_fuel_four_columns(headers)
        if m:
            ci, di, qi, ai = m
            return hr, ci, di, qi, ai
    return None


def _toll_col_key(name) -> str:
    """Ustun nomini taqqoslash uchun kalit (faqat harf va raqam)."""
    import re

    flat = _flatten_column_name(name).replace("\ufeff", "")
    s = re.sub(r"[^a-z0-9]", "", flat.lower())
    return s


def match_toll_named_columns(df) -> tuple[str, str, str] | None:
    """
    PostingDate, PPTagID, Toll_Amount ustunlarini topadi.
    Qaytaradi: (posting_col, pptag_col, toll_amount_col) yoki None.
    """
    import pandas as pd

    if df is None or df.empty:
        return None
    col_map = {}
    for c in df.columns:
        k = _toll_col_key(c)
        if k and k not in col_map:
            col_map[k] = c

    def get_col(*aliases):
        for a in aliases:
            ak = _toll_col_key(a)
            if ak in col_map:
                return col_map[ak]
        return None

    posting = get_col("PostingDate", "postingdate", "post date", "postdate")
    if not posting:
        for k, col in col_map.items():
            if "posting" in k and "date" in k:
                posting = col
                break
    if not posting:
        for k, col in col_map.items():
            if k in ("postdate",) or (k.startswith("post") and k.endswith("date") and "invoice" not in k):
                posting = col
                break

    pptag = get_col("PPTagID", "pptagid", "pptag", "ppdeviceid", "ppdevice")
    if not pptag:
        for k, col in col_map.items():
            if "pptag" in k:
                pptag = col
                break

    toll_amt = get_col("Toll_Amount", "tollamount", "toll_amt", "toll amt")
    if not toll_amt:
        for k, col in col_map.items():
            if "toll" in k and "amount" in k:
                toll_amt = col
                break

    if posting and pptag and toll_amt:
        return posting, pptag, toll_amt
    return None


def parse_toll_posting_date(val):
    """PostingDate (masalan DD.MM.YYYY) -> date yoki None."""
    import re
    from datetime import datetime

    import pandas as pd

    if val is None or (isinstance(val, float) and val != val):
        return None
    if pd.isna(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val.date()
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    if "T" in s:
        s = s.split("T", 1)[0].strip()
    elif " " in s and re.match(r"^\d{4}-\d{2}-\d{2}", s):
        s = s.split()[0]
    elif " " in s and re.match(r"^\d{1,2}\.\d{1,2}\.\d{4}", s):
        s = s.split()[0]
    # Excel serial (raqam yoki "45452" qatori)
    try:
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            v = float(val)
            if 30000 < v < 60000:
                from datetime import date, timedelta

                base = date(1899, 12, 30)
                return base + timedelta(days=int(v))
    except Exception:
        pass
    try:
        if re.match(r"^\d+\.?\d*$", s):
            v = float(s)
            if 30000 < v < 60000:
                from datetime import date, timedelta

                base = date(1899, 12, 30)
                return base + timedelta(days=int(v))
    except Exception:
        pass
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    try:
        return pd.to_datetime(val, dayfirst=True).date()
    except Exception:
        return None


def expense_item_date_in_segment(item_date, seg_start, seg_end) -> bool:
    """
    Exceldagi tranzaksiya sanasi (yil farq qilishi mumkin) sheetdagi hafta oralig'iga tushadimi.
    Masalan: sheet "2024 ... 03.30-04.05", Excelda 01.04.2026 — oy-kun segment yiliga
    moslashtirilib tekshiriladi (mart-aprel oralig'idagi aprel kunlari ham kiradi).
    """
    from datetime import date as date_cls

    years_to_try = {seg_start.year, seg_end.year, item_date.year}
    years_to_try.add(seg_start.year - 1)
    years_to_try.add(seg_end.year + 1)

    for y in sorted(years_to_try):
        try:
            d = date_cls(y, item_date.month, item_date.day)
        except ValueError:
            if item_date.month == 2 and item_date.day == 29:
                try:
                    d = date_cls(y, 2, 28)
                except ValueError:
                    continue
            else:
                continue
        if seg_start <= seg_end:
            if seg_start <= d <= seg_end:
                return True
        else:
            if d >= seg_start or d <= seg_end:
                return True
    return False


def parse_toll_amount_positive_only(val):
    """
    Toll summasini qaytaradi; manfiy yoki '-' bilan boshlangan qiymatlar None (tashlab ketiladi).
    Vergul bilan o'nlik ajratuvchi (0,9) qo'llab-quvvatlanadi.
    """
    import re

    import pandas as pd

    if val is None or (isinstance(val, float) and val != val):
        return None
    if pd.isna(val):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        if val < 0:
            return None
        return float(val)
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    s_clean = s.replace("$", "").replace(" ", "").replace("\xa0", "")
    if s_clean.startswith("-"):
        return None
    if "," in s_clean and "." in s_clean:
        if s_clean.rfind(",") > s_clean.rfind("."):
            s_clean = s_clean.replace(".", "").replace(",", ".")
        else:
            s_clean = s_clean.replace(",", "")
    elif "," in s_clean and "." not in s_clean:
        parts = s_clean.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            s_clean = parts[0].replace(".", "") + "." + parts[1]
        else:
            s_clean = s_clean.replace(",", "")
    else:
        s_clean = s_clean.replace(",", "")
    m = re.search(r"\d+(\.\d+)?", s_clean)
    if not m:
        return None
    try:
        f = float(m.group())
        if f < 0:
            return None
        return f
    except ValueError:
        return None


def find_fuel_columns_from_named_dataframe(df) -> tuple[int, int, int, int] | None:
    """pd.read_excel(header=0) — ustun nomlari Card #, Tran Date, ... bo'lsa."""
    if df is None or df.empty:
        return None
    headers = []
    for c in df.columns:
        headers.append(_fuel_norm_header(_flatten_column_name(c)))
    return match_fuel_four_columns(headers)


def autopick_fuel_expense_tab(sheet_candidates: list, fuel_entries: list) -> str | None:
    """Barcha tranzaksiya sanalari bitta hafta-list oralig'iga tushsa, o'sha list nomini qaytaradi."""
    import re
    from datetime import datetime, date
    if not sheet_candidates or not fuel_entries:
        return None
    dates: list = []
    for e in fuel_entries:
        try:
            dates.append(datetime.fromisoformat(e["date"]).date())
        except Exception:
            continue
    if not dates:
        return None
    min_d, max_d = min(dates), max(dates)
    range_re = re.compile(r"(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})")
    fits: list[str] = []
    for name in sheet_candidates:
        m = range_re.search(name)
        if not m:
            continue
        year_m = re.search(r"(\d{4})", name)
        year = int(year_m.group(1)) if year_m else date.today().year
        start_str, end_str = m.group(1), m.group(2)
        sm, sd = map(int, start_str.split("."))
        em, ed = map(int, end_str.split("."))
        start_date = date(year, sm, sd)
        end_year = year
        if (em, ed) < (sm, sd):
            end_year = year + 1
        end_date = date(end_year, em, ed)
        if start_date <= min_d and max_d <= end_date:
            fits.append(name)
    if len(fits) == 1:
        return fits[0]
    return None


async def apply_fuel_named_week_to_sheet(
    status_msg,
    state: FSMContext,
    company: str,
    sheet_name: str,
    fuel_entries: list,
    fuel_filename: str,
) -> bool:
    """
    Google Sheet varag'i nomida MM.DD-MM.DD oralig'i bo'lsa — yozuv + Excel hisobot.
    False: bu varaq uslubida ishlamaydi (keyingi filialga o'ting).
    """
    import os
    import re
    import tempfile
    import pandas as pd
    from datetime import datetime, date
    from aiogram.types import FSInputFile
    from services.google_sheets import get_sheet_service
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    range_re = re.compile(r"(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})")
    year_m = re.search(r"(\d{4})", sheet_name)
    year = int(year_m.group(1)) if year_m else datetime.now().year
    m = range_re.search(sheet_name)
    if not m:
        return False

    start_str, end_str = m.group(1), m.group(2)
    start_month, start_day = map(int, start_str.split("."))
    end_month, end_day = map(int, end_str.split("."))
    start_date = date(year, start_month, start_day)
    end_year = year
    if (end_month, end_day) < (start_month, start_day):
        end_year = year + 1
    end_date = date(end_year, end_month, end_day)

    sheet_service = get_sheet_service()
    card_totals = {}
    for item in fuel_entries:
        try:
            item_date = datetime.fromisoformat(item["date"]).date()
        except Exception:
            continue
        if not (start_date <= item_date <= end_date):
            continue
        card = str(item.get("card", "")).strip()
        if not card:
            continue
        fuel_sum = float(item.get("fuel", 0.0) or 0.0)
        discount_sum = float(item.get("discount", 0.0) or 0.0)
        if card not in card_totals:
            card_totals[card] = [0.0, 0.0]
        card_totals[card][0] += fuel_sum
        card_totals[card][1] += discount_sum

    if not card_totals:
        await status_msg.edit_text(f"List <b>{sheet_name}</b> oralig'ida mos yozuv topilmadi.")
        await state.set_state(BotStates.Fuel)
        return True

    card_totals_tuple = {k: (v[0], v[1]) for k, v in card_totals.items()}
    updated, skipped, missing_count, missing_cards = sheet_service.update_fuel_toll_expenses(
        sheet_name,
        card_totals_tuple,
        fuel_col=5,
        discount_col=6,
        company=company,
    )

    report_rows = []
    missing_set = set(str(x) for x in (missing_cards or []))
    week_label = f"{start_str}-{end_str}"
    for card, (fuel_sum, discount_sum) in card_totals_tuple.items():
        card_s = str(card)
        if (fuel_sum or 0) == 0 and (discount_sum or 0) == 0:
            continue
        status = "TOPILMADI" if card_s in missing_set else "TOPILDI"
        report_rows.append(
            {
                "Sheet": sheet_name,
                "Week": week_label,
                "Card": card_s,
                "FuelSum": fuel_sum,
                "DiscountSum": discount_sum,
                "Status": status,
            }
        )

    try:
        report_df = pd.DataFrame(report_rows)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = tmp.name
        tmp.close()
        report_df.to_excel(tmp_path, index=False)
        try:
            wb = load_workbook(tmp_path)
            ws_rep = wb.active
            status_col_idx = None
            for c in range(1, ws_rep.max_column + 1):
                if str(ws_rep.cell(row=1, column=c).value).strip() == "Status":
                    status_col_idx = c
                    break
            if status_col_idx:
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for r in range(2, ws_rep.max_row + 1):
                    cell = ws_rep.cell(row=r, column=status_col_idx)
                    v = str(cell.value).strip().upper() if cell.value is not None else ""
                    if v == "TOPILDI":
                        cell.fill = green_fill
                    elif v == "TOPILMADI":
                        cell.fill = red_fill
                wb.save(tmp_path)
        except Exception:
            pass
        await status_msg.answer_document(FSInputFile(tmp_path))
        os.remove(tmp_path)
    except Exception:
        pass

    await status_msg.edit_text("✅ Fuel yozildi. (Excel report yuborildi)")
    await state.set_state(BotStates.Fuel)
    return True


@dp.message(F.text == "⛽ Fuel Expenses")
async def enter_fuel(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    await state.set_state(BotStates.Fuel)
    await message.answer(
        "Excel (xlsx, xls) yuboring. Jadvalda <b>Card #</b>, <b>Tran Date</b>, <b>Disc Amt</b>, <b>Amt</b> "
        "sarlavhalari bo'lsa — bot ustunlarni o'zi topadi; hafta ro'yxati aniq bo'lsa, Google Sheetga avtomatik yozadi.",
        reply_markup=expenses_menu,
    )

@dp.message(F.text == "🛣️ Toll Expenses")
async def enter_toll(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    await state.set_state(BotStates.Toll)
    await message.answer(
        "Toll Expenses bo'limi.\n"
        "Excel (xlsx, xls) yuboring — jadvalda <b>PostingDate</b>, <b>PPTagID</b>, <b>Toll_Amount</b> "
        "ustunlari bo'lsa bot o'zi topadi (manfiy Toll_Amount qatorlari hisobga olinmaydi).",
        reply_markup=expenses_menu,
    )

@dp.message(F.text == "⬅️ Back (Main Menu)", BotStates.Fuel)
@dp.message(F.text == "⬅️ Back (Main Menu)", BotStates.Toll)
async def back_main(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Asosiy menyu:", reply_markup=get_main_menu(message.from_user.id))

@dp.message(F.document, BotStates.Fuel)
async def handle_fuel_doc(message: types.Message, state: FSMContext):
    await handle_expense_doc(message, "FUEL", state)

@dp.message(F.document, BotStates.Toll)
async def handle_toll_doc(message: types.Message, state: FSMContext):
    await handle_expense_doc(message, "TOLL", state)

async def handle_expense_doc(message: types.Message, expense_type: str, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    document = message.document
    file_id = document.file_id
    file_name = document.file_name

    if expense_type == "FUEL" and not ("xlsx" in file_name.lower() or "xls" in file_name.lower()):
        await message.answer("⚠️ Fuel uchun faqat Excel (xlsx, xls) fayl yuboring.")
        return

    if expense_type == "TOLL" and not ("xlsx" in file_name.lower() or "xls" in file_name.lower()):
        await message.answer("⚠️ Toll uchun faqat Excel (xlsx, xls) fayl yuboring.")
        return

    await message.answer(f"{expense_type} fayl qabul qilindi. Qayta ishlanmoqda... ⏳")

    try:
        file = await bot.get_file(file_id)
        file_content = await bot.download_file(file.file_path)
        content_bytes = file_content.read()

        from services.google_sheets import get_sheet_service
        sheet_service = get_sheet_service()

        import pandas as pd
        import io
        import re

        # -------------------- FUEL --------------------
        if expense_type == "FUEL":
            # Ustunlar sarlavha bo'yicha topiladi: Card #, Tran Date, Disc Amt, Amt (tartib ixtiyoriy).
            # Bir nechta varaq bo'lsa, birinchisi bo'yicha ma'lumot chiqadigan varaq tanlanadi.

            def parse_money(val):
                if val is None or (isinstance(val, float) and val != val):
                    return 0.0
                if pd.isna(val):
                    return 0.0
                s = str(val).strip().replace("$", "").replace(" ", "")
                # Decimal vergul bo'lishi mumkin: "2,13" => "2.13"
                if "," in s and "." in s:
                    s = s.replace(",", "")
                elif "," in s and "." not in s:
                    s = s.replace(",", ".")
                m = re.search(r'-?\d+(\.\d+)?', s)
                if not m:
                    return 0.0
                try:
                    return float(m.group())
                except ValueError:
                    return 0.0

            def normalize_card_str(v: str) -> str:
                if v is None:
                    return ""
                s = str(v).strip()
                if not s or s.lower() == "nan":
                    return ""
                # ilmiy yozuv bo'lsa
                if "e" in s.lower():
                    try:
                        from decimal import Decimal
                        return str(int(Decimal(s)))
                    except Exception:
                        pass
                # trailing .0
                if s.endswith(".0"):
                    return s[:-2]
                return s

            def parse_fuel_tran_date(val):
                """
                EFS/Transaction report: B ustun — odatda DD.MM.YYYY (Yevropa).
                pd.to_datetime() defaultda 01.04.2026 ni AQSH sifatida (4-yanvar) o'qiydi — noto'g'ri.
                Excel ba'zan sanani serial raqam (float) qilib beradi.
                """
                import datetime as dt_mod

                if val is None or (isinstance(val, float) and val != val):
                    return None
                if pd.isna(val):
                    return None
                if isinstance(val, pd.Timestamp):
                    return val.date()
                if isinstance(val, dt_mod.datetime):
                    return val.date()
                if isinstance(val, dt_mod.date):
                    return val
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    try:
                        n = float(val)
                        if 20000 < n < 100000:
                            dt = pd.to_datetime(n, unit="D", origin="1899-12-30", errors="coerce")
                            if pd.notna(dt):
                                return dt.date()
                    except Exception:
                        pass
                s = str(val).strip()
                if not s or s.lower() == "nan":
                    return None
                try:
                    dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
                    if pd.isna(dt):
                        return None
                    return dt.date()
                except Exception:
                    return None

            def fuel_rows_to_acc(df_part, ci, di, qi, ai):
                """df_part — faqat ma'lumot qatorlari (sarlavhasiz)."""
                acc = {}
                for r in range(len(df_part)):
                    row = df_part.iloc[r]
                    n = len(row)
                    def gc(j):
                        return row.iloc[j] if j < n else None
                    card_val = normalize_card_str(gc(ci))
                    if not card_val:
                        continue
                    trans_date = parse_fuel_tran_date(gc(di))
                    if trans_date is None:
                        continue
                    date_iso = trans_date.isoformat()
                    discount_sum = parse_money(gc(qi))
                    fuel_sum = parse_money(gc(ai))
                    key = (card_val, date_iso)
                    if key not in acc:
                        acc[key] = [0.0, 0.0]
                    acc[key][0] += fuel_sum
                    acc[key][1] += discount_sum
                return acc

            try:
                xls = pd.ExcelFile(io.BytesIO(content_bytes))
            except Exception:
                await message.answer("❌ Fuel uchun xlsx faylni o‘qib bo‘lmadi.")
                return

            last_progress = await message.answer("⏳ Fuel xlsx o‘qilmoqda...")

            entries_acc = {}
            sheets_tried = []
            for sheet_name in xls.sheet_names:
                sheets_tried.append(sheet_name)
                acc = {}

                # A) Birinchi qator = ustun nomlari (Delo / EFS eksport — eng ko'p holat)
                try:
                    df0 = pd.read_excel(xls, sheet_name=sheet_name, header=0, dtype=object)
                except Exception:
                    df0 = None
                if df0 is not None and df0.shape[1] >= 4:
                    colmap = find_fuel_columns_from_named_dataframe(df0)
                    if colmap:
                        ci, di, qi, ai = colmap
                        acc = fuel_rows_to_acc(df0, ci, di, qi, ai)

                if acc:
                    entries_acc = acc
                    break

                # B) Sarlavha 2–45 qatorlarda yoki birlashtirilgan kataklar tufayli header=None
                try:
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=object)
                except Exception:
                    continue
                if df.shape[1] < 4:
                    continue
                found = find_fuel_transaction_header_map(df)
                if found:
                    hr, ci, di, qi, ai = found
                    acc = fuel_rows_to_acc(df.iloc[hr + 1 :], ci, di, qi, ai)
                if acc:
                    entries_acc = acc
                    break

            if not entries_acc:
                hint = ""
                if not sheets_tried:
                    hint = "Fayl ichida varaq topilmadi."
                else:
                    hint = (
                        "Kerakli ustunlar topilmadi yoki yozuvlar o'qilmadi (sana/karta). "
                        "Birinchi qatorda <b>Card #</b>, <b>Tran Date</b>, <b>Disc Amt</b>, <b>Amt</b> bo'lsin; "
                        "faylni Excelda <b>.xlsx</b> qilib qayta saqlang (CSV emas)."
                    )
                try:
                    await last_progress.edit_text(
                        "❌ Fuel fayldan ma'lumot chiqmadi.\n" + hint
                    )
                except Exception:
                    await message.answer("❌ Fuel fayldan ma'lumot chiqmadi.\n" + hint)
                return

            try:
                await last_progress.edit_text("✅ Fayl tayyor. Qaysi listni tekshiramiz?")
            except Exception:
                pass

            # Sheets listlarini ko‘rsatamiz (expenses spreadsheetdan)
            expenses_sheet_names = sheet_service.get_expenses_all_sheet_names(company)
            range_re = re.compile(r'(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})')
            sheet_candidates = [s for s in expenses_sheet_names if range_re.search(s)]
            if not sheet_candidates:
                sheet_candidates = expenses_sheet_names

            # Statega saqlaymiz (discount = Disc Amt, Discount ustuniga yoziladi)
            fuel_entries = []
            for (card, date_iso), (fuel_sum, discount_sum) in entries_acc.items():
                fuel_entries.append({"card": card, "date": date_iso, "fuel": fuel_sum, "discount": discount_sum})

            await state.set_state(BotStates.FuelSheetSelect)
            await state.update_data(
                fuel_entries=fuel_entries,
                fuel_sheet_names=sheet_candidates,
                fuel_filename=file_name,
                selected_company=company,
            )

            picked_tab = autopick_fuel_expense_tab(sheet_candidates, fuel_entries)
            if picked_tab:
                try:
                    await last_progress.edit_text(
                        f"✅ Ro'yxat: <b>{picked_tab}</b> (avtomatik). Google Sheets ga yozilmoqda..."
                    )
                except Exception:
                    pass
                await apply_fuel_named_week_to_sheet(
                    last_progress, state, company, picked_tab, fuel_entries, file_name
                )
                return

            # Inline keyboard (2 ustun)
            buttons = []
            row = []
            for i, name in enumerate(sheet_candidates):
                row.append(InlineKeyboardButton(text=name, callback_data=f"fuel_sheet:{i}"))
                if len(row) == 2:
                    buttons.append(row)
                    row = []
            if row:
                buttons.append(row)
            kb = InlineKeyboardMarkup(inline_keyboard=buttons)

            await message.answer(
                "📋 Qaysi listni tekshiramiz?\n"
                "Tugmani bosing — natija tez orada chiqadi.",
                reply_markup=kb
            )
            return

        # -------------------- TOLL (PostingDate / PPTagID / Toll_Amount yoki PrePass legacy) --------------------
        if expense_type == "TOLL":
            df = None
            toll_mode = None  # "named" | "prepass"
            try:
                df = pd.read_excel(io.BytesIO(content_bytes), sheet_name=0, header=0, dtype=str)
            except Exception:
                df = None

            named = match_toll_named_columns(df) if df is not None else None

            if named:
                toll_mode = "named"
                col_post, col_pp, col_toll = named
            else:
                toll_mode = None
                try:
                    df = pd.read_excel(io.BytesIO(content_bytes), sheet_name=0, header=8, dtype=str)
                except Exception:
                    try:
                        df = pd.read_excel(io.BytesIO(content_bytes), sheet_name=0, dtype=str)
                    except Exception:
                        await message.answer("❌ Toll uchun xlsx faylni o'qib bo'lmadi.")
                        return
                if df is not None and df.shape[1] > 22:
                    toll_mode = "prepass"
                else:
                    await message.answer(
                        "❌ Toll faylda PostingDate, PPTagID, Toll_Amount ustunlari topilmadi "
                        "(yoki PrePass uchun kamida 23 ustun kerak)."
                    )
                    return

            def normalize_transponder(v):
                if v is None or pd.isna(v):
                    return ""
                s = str(v).strip()
                if not s or s.lower() == "nan":
                    return ""
                if s.endswith(".0"):
                    s = s[:-2]
                return s

            def parse_money_legacy(val):
                if val is None or (isinstance(val, float) and val != val) or pd.isna(val):
                    return 0.0
                s = str(val).strip().replace("$", "").replace(" ", "").replace(",", "")
                m = re.search(r"-?\d+(\.\d+)?", s)
                if not m:
                    return 0.0
                try:
                    return float(m.group())
                except ValueError:
                    return 0.0

            entries_acc = {}

            if toll_mode == "named":
                for _, row in df.iterrows():
                    date_val = row[col_post]
                    pp_id = normalize_transponder(row[col_pp])
                    toll_val = parse_toll_amount_positive_only(row[col_toll])

                    if not pp_id or toll_val is None:
                        continue
                    trans_date = parse_toll_posting_date(date_val)
                    if not trans_date:
                        continue

                    key = (pp_id, trans_date.isoformat())
                    if key not in entries_acc:
                        entries_acc[key] = 0.0
                    entries_acc[key] += toll_val

            else:
                # PrePass: A=Post Date, E=PP Device ID, W=Toll $
                for _, row in df.iterrows():
                    date_val = row.iloc[0]
                    pp_id = normalize_transponder(row.iloc[4])
                    toll_raw = parse_money_legacy(row.iloc[22])
                    if toll_raw < 0:
                        continue
                    if not pp_id:
                        continue
                    try:
                        trans_date = pd.to_datetime(date_val).date()
                    except Exception:
                        continue

                    key = (pp_id, trans_date.isoformat())
                    if key not in entries_acc:
                        entries_acc[key] = 0.0
                    entries_acc[key] += toll_raw

            toll_entries = [
                {"transponder": pp_id, "date": date_iso, "toll": toll_sum}
                for (pp_id, date_iso), toll_sum in entries_acc.items()
            ]

            if not toll_entries:
                await message.answer(
                    "❌ Toll faylda ma'lumot topilmadi (PostingDate, PPTagID, Toll_Amount "
                    "yoki PrePass A/E/W tekshirilsin; manfiy Toll_Amount qatorlari hisobga olinmaydi)."
                )
                return

            expenses_sheet_names = sheet_service.get_expenses_all_sheet_names(company)
            range_re = re.compile(r'(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})')
            sheet_candidates = [s for s in expenses_sheet_names if range_re.search(s)]
            if not sheet_candidates:
                sheet_candidates = expenses_sheet_names

            await state.set_state(BotStates.TollSheetSelect)
            await state.update_data(
                toll_entries=toll_entries,
                toll_sheet_names=sheet_candidates,
                toll_filename=file_name,
                selected_company=company,
            )

            buttons = []
            row_btns = []
            for i, name in enumerate(sheet_candidates):
                row_btns.append(InlineKeyboardButton(text=name, callback_data=f"toll_sheet:{i}"))
                if len(row_btns) == 2:
                    buttons.append(row_btns)
                    row_btns = []
            if row_btns:
                buttons.append(row_btns)
            kb = InlineKeyboardMarkup(inline_keyboard=buttons)

            await message.answer(
                "✅ Fayl tayyor. Qaysi listni tekshiramiz?\n"
                "Tugmani bosing — natija tez orada chiqadi.",
                reply_markup=kb
            )
            return

        await message.answer("❌ Noma'lum expense_type.")

    except Exception as e:
        await message.answer(f"❌ Xatolik yuz berdi: {e}")


@dp.callback_query(F.data.startswith("fuel_sheet:"), BotStates.FuelSheetSelect)
async def callback_fuel_sheet(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()

    data = await state.get_data()
    company = data.get("selected_company") or get_company(callback.from_user.id)
    if not company:
        await callback.message.edit_text("Iltimos, /start bosing va Load tanlang.")
        return
    sheet_names = data.get("fuel_sheet_names") or []
    fuel_entries = data.get("fuel_entries") or []
    if not sheet_names or not fuel_entries:
        await callback.message.edit_text(
            "❌ Ma'lumotlar topilmadi (sessiya tugagan yoki faylda yozuv yo'q). "
            "Qaytadan Fuel fayl yuboring — bot qayta ishga tushgan bo'lsa, state saqlanmaydi."
        )
        await state.set_state(BotStates.Fuel)
        return

    idx_str = callback.data.replace("fuel_sheet:", "").strip()
    try:
        idx = int(idx_str)
    except ValueError:
        return
    if idx < 0 or idx >= len(sheet_names):
        return

    sheet_name = sheet_names[idx]

    await callback.message.edit_text("⏳ Kutib turing, natija tez orada chiqadi...")

    import re
    from datetime import datetime, date
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    from services.google_sheets import get_sheet_service

    sheet_service = get_sheet_service()

    year_m = re.search(r"(\d{4})", sheet_name)
    year = int(year_m.group(1)) if year_m else datetime.now().year

    # 1) Sheet nomida hafta oralig'i (masalan 03.30-04.05) bo'lsa — umumiy yozuv funksiyasi
    range_re = re.compile(r'(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})')
    if range_re.search(sheet_name):
        fn = data.get("fuel_filename") or "fuel.xlsx"
        await apply_fuel_named_week_to_sheet(
            callback.message, state, company, sheet_name, fuel_entries, fn
        )
        return

    # 2) Agar sheet nomida oralig' bo'lmasa (masalan Owner Operators) -> sheet ichidan sana va ustunlarni topamiz
    ws = sheet_service.get_expenses_board(sheet_name, company)
    if not ws:
        await callback.message.edit_text(f"❌ {sheet_name} worksheet topilmadi.")
        await state.set_state(BotStates.Fuel)
        return

    # Faqat yuqori qismi: sana oralig'i va Fuel/Toll header joylari topiladi
    top_grid = ws.get("A1:Z6")
    # top_grid: list rows, col count = Z

    seg_re = re.compile(r'(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})')
    date_matches = []  # (start_date, end_date, start_col_idx)

    # sana oralig'i odatda row 1-3 da bo'ladi
    for r in range(min(len(top_grid), 10)):
        for c in range(min(26, len(top_grid[r]))):
            cell = top_grid[r][c]
            if cell is None:
                continue
            cell_s = str(cell)
            m2 = seg_re.search(cell_s)
            if not m2:
                continue
            # cell ichida yil bo'lishi mumkin (masalan: "2026 February/March 02.23-03.01")
            year_cell_m = re.search(r'(\d{4})', cell_s)
            year_cell = int(year_cell_m.group(1)) if year_cell_m else year
            start_str, end_str = m2.group(1), m2.group(2)
            start_month, start_day = map(int, start_str.split('.'))
            end_month, end_day = map(int, end_str.split('.'))

            s_date = date(year_cell, start_month, start_day)
            e_year = year_cell
            if (end_month, end_day) < (start_month, start_day):
                e_year = year_cell + 1
            e_date = date(e_year, end_month, end_day)
            # col index 0-based -> 1-based for gspread usage
            date_matches.append((s_date, e_date, c + 1))

    # Unikallashtirish
    unique = {}
    for s_date, e_date, c in date_matches:
        key = (s_date, e_date, c)
        unique[key] = True
    date_matches = sorted(list(unique.keys()), key=lambda x: x[2])

    # Card/Fuel/Discount ustun topish (Disc Amt -> Discount, Toll Exp emas)
    card_cols = []
    fuel_cols = []
    discount_cols = []
    # Taxmin: Fuel/Discount header row 1-6 atrofida
    for r in range(min(6, len(top_grid))):
        for c in range(min(26, len(top_grid[r]))):
            cell = top_grid[r][c]
            if cell is None:
                continue
            cell_s = str(cell).lower()

            # "Fuel after discount", "Fuel Exp", "Fuel Expenses" kabi
            if "fuel" in cell_s and ("exp" in cell_s or "after" in cell_s):
                fuel_cols.append(c + 1)

            # "EFS Card" ustuni (Owner Operators: C, Company Drivers: D)
            if "efs" in cell_s and "card" in cell_s:
                card_cols.append(c + 1)

            # "Discount", "Disc Amt" - Disc Amt (Q) shu ustunga yoziladi
            # "Fuel after discount" ni hisobga olmaslik: fuel bo'lmagan discount
            if "fuel" not in cell_s and ("discount" in cell_s or "disc" in cell_s):
                discount_cols.append(c + 1)

    # Agar topilmasa, fallback: E=Fuel, F=Discount
    if not fuel_cols:
        fuel_cols = [5]
    if not discount_cols:
        discount_cols = [6]
    if not card_cols:
        # Eski layout fallback: C ustun
        card_cols = [3]

    # Har bir sana oralig'i uchun fuel/discount colni segmentga moslab tanlaymiz.
    # Segment: date_matches[i].start_col -> date_matches[i+1].start_col-1
    segments = []
    if date_matches:
        for i in range(len(date_matches)):
            start_date, end_date, start_col = date_matches[i]
            end_col = (date_matches[i + 1][2] - 1) if i + 1 < len(date_matches) else 26
            # segment ichida joylashgan fuel/discount header col'larini tanlaymiz
            f_col = next((fc for fc in fuel_cols if start_col <= fc <= end_col), None)
            d_col = next((dc for dc in discount_cols if start_col <= dc <= end_col), None)
            c_col = next((cc for cc in card_cols if start_col <= cc <= end_col), None)
            if f_col is None:
                f_col = fuel_cols[0]
            if d_col is None:
                d_col = discount_cols[0]
            if c_col is None:
                c_col = card_cols[0]
            segments.append({
                "label": f"{start_date.strftime('%m.%d')}-{end_date.strftime('%m.%d')}",
                "start_date": start_date,
                "end_date": end_date,
                "card_col": c_col,
                "fuel_col": f_col,
                "discount_col": d_col,
            })
    else:
        # sana topilmasa - bitta umumiy segment
        segments = [{
            "label": sheet_name,
            "start_date": date(year, 1, 1),
            "end_date": date(year, 12, 31),
            "card_col": card_cols[0],
            "fuel_col": fuel_cols[0],
            "discount_col": discount_cols[0],
        }]

    # Auto-mapping: foydalanuvchi sana oralig'ini tanlamaydi,
    # xlsxdagi har bir sana o'z segmentiga tushib, shu haftaga yoziladi.
    card_totals_by_seg = {i: {} for i in range(len(segments))}
    matched_items = 0

    for item in fuel_entries:
        try:
            item_date = datetime.fromisoformat(item["date"]).date()
        except Exception:
            continue

        card = str(item.get("card", "")).strip()
        if not card:
            continue

        fuel_sum = float(item.get("fuel", 0.0) or 0.0)
        discount_sum = float(item.get("discount", 0.0) or 0.0)

        assigned = False
        for i, seg in enumerate(segments):
            if not expense_item_date_in_segment(item_date, seg["start_date"], seg["end_date"]):
                continue

            if card not in card_totals_by_seg[i]:
                card_totals_by_seg[i][card] = [0.0, 0.0]
            card_totals_by_seg[i][card][0] += fuel_sum
            card_totals_by_seg[i][card][1] += discount_sum
            assigned = True
            break

        if assigned:
            matched_items += 1

    total_updated = 0
    total_skipped = 0
    missing_cards_by_seg = {}

    for i, seg in enumerate(segments):
        card_totals = card_totals_by_seg.get(i) or {}
        if not card_totals:
            continue

        card_totals_tuple = {k: (v[0], v[1]) for k, v in card_totals.items()}
        updated, skipped, missing_count, missing_cards = sheet_service.update_fuel_toll_expenses(
            sheet_name,
            card_totals_tuple,
            fuel_col=seg["fuel_col"],
            discount_col=seg["discount_col"],
            card_col=seg.get("card_col", 3),
            company=company,
        )
        total_updated += updated
        total_skipped += skipped
        missing_cards_by_seg[i] = set(str(x) for x in (missing_cards or []))

    if matched_items == 0:
        await callback.message.edit_text(
            f"❌ {sheet_name} bo'yicha mos hafta topilmadi.\n\n"
            f"Fayldagi sana oralig'i: xlsx -> (tekshirilsin)."
        )
        await state.set_state(BotStates.Fuel)
        return

    # Excel report (TOPILDI/TOPILMADI va qaysi hafta)
    report_rows = []
    for i, seg in enumerate(segments):
        seg_cards = card_totals_by_seg.get(i) or {}
        if not seg_cards:
            continue
        seg_missing = missing_cards_by_seg.get(i) or set()
        for card, (fuel_sum, discount_sum) in seg_cards.items():
            card_s = str(card)
            if (fuel_sum or 0) == 0 and (discount_sum or 0) == 0:
                continue
            status = "TOPILMADI" if card_s in seg_missing else "TOPILDI"
            report_rows.append({
                "Sheet": sheet_name,
                "Week": seg.get("label"),
                "Card": card_s,
                "FuelSum": fuel_sum,
                "DiscountSum": discount_sum,
                "Status": status,
            })

    try:
        import pandas as pd
        import os
        import tempfile
        from aiogram.types import FSInputFile

        report_df = pd.DataFrame(report_rows)
        data_all = await state.get_data()
        src_file = data_all.get("fuel_filename") or "fuel.xlsx"
        base = os.path.splitext(str(src_file))[0]

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = tmp.name
        tmp.close()

        report_df.to_excel(tmp_path, index=False)
        # Status ustunini bo'yash: TOPILDI=green, TOPILMADI=red
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            wb = load_workbook(tmp_path)
            ws_rep = wb.active
            status_col_idx = None
            for c in range(1, ws_rep.max_column + 1):
                if str(ws_rep.cell(row=1, column=c).value).strip() == "Status":
                    status_col_idx = c
                    break

            if status_col_idx:
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for r in range(2, ws_rep.max_row + 1):
                    cell = ws_rep.cell(row=r, column=status_col_idx)
                    v = str(cell.value).strip().upper() if cell.value is not None else ""
                    if v == "TOPILDI":
                        cell.fill = green_fill
                    elif v == "TOPILMADI":
                        cell.fill = red_fill

                wb.save(tmp_path)
        except Exception:
            pass
        await callback.message.answer_document(FSInputFile(tmp_path))
        os.remove(tmp_path)
    except Exception:
        pass

    await callback.message.edit_text("✅ Fuel yozildi. (Excel report yuborildi)")

    await state.set_state(BotStates.Fuel)


@dp.callback_query(F.data.startswith("fuel_range:"), BotStates.FuelRangeSelect)
async def callback_fuel_range(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    from datetime import datetime
    data = await state.get_data()
    company = data.get("selected_company") or get_company(callback.from_user.id)
    if not company:
        await callback.message.edit_text("Iltimos, /start bosing va Load tanlang.")
        return
    segments = data.get("fuel_segments") or []
    fuel_entries = data.get("fuel_entries") or []
    if not segments or not fuel_entries:
        await callback.message.edit_text("❌ Ma'lumotlar yo'q. Qaytadan Fuel fayl yuboring.")
        await state.set_state(BotStates.Fuel)
        return

    idx_str = callback.data.replace("fuel_range:", "").strip()
    try:
        idx = int(idx_str)
    except ValueError:
        return
    if idx < 0 or idx >= len(segments):
        return

    seg = segments[idx]
    sheet_name = data.get("fuel_selected_sheet") or ""
    if not sheet_name:
        await callback.message.edit_text("❌ Tanlangan sheet nomi topilmadi. Qaytadan ishlang.")
        await state.set_state(BotStates.Fuel)
        return

    await callback.message.edit_text("⏳ Kutib turing, natija tez orada chiqadi...")

    # Filter faqat shu oralig'idagi yozuvlar (sheet yili bilan oy-kun moslashtiriladi).
    card_totals = {}

    for item in fuel_entries:
        try:
            item_date = datetime.fromisoformat(item["date"]).date()
        except Exception:
            continue
        if not expense_item_date_in_segment(item_date, seg["start_date"], seg["end_date"]):
            continue
        card = str(item.get("card", "")).strip()
        if not card:
            continue
        fuel_sum = float(item.get("fuel", 0.0) or 0.0)
        discount_sum = float(item.get("discount", 0.0) or 0.0)
        if card not in card_totals:
            card_totals[card] = [0.0, 0.0]
        card_totals[card][0] += fuel_sum
        card_totals[card][1] += discount_sum

    if not card_totals:
        await callback.message.edit_text(f"List <b>{sheet_name}</b> oralig'ida mos yozuv topilmadi.")
        await state.set_state(BotStates.Fuel)
        return

    from services.google_sheets import get_sheet_service
    sheet_service = get_sheet_service()

    card_totals_tuple = {k: (v[0], v[1]) for k, v in card_totals.items()}
    updated, skipped, missing_count, missing_cards = sheet_service.update_fuel_toll_expenses(
        sheet_name,
        card_totals_tuple,
        fuel_col=seg["fuel_col"],
        discount_col=seg["discount_col"],
        company=company,
    )

    report_rows = []
    missing_set = set(str(x) for x in (missing_cards or []))
    week_label = seg.get("label")
    for card, (fuel_sum, discount_sum) in card_totals_tuple.items():  # fuel_range
        card_s = str(card)
        if (fuel_sum or 0) == 0 and (discount_sum or 0) == 0:
            continue
        status = "TOPILMADI" if card_s in missing_set else "TOPILDI"
        report_rows.append({
            "Sheet": sheet_name,
            "Week": week_label,
            "Card": card_s,
            "FuelSum": fuel_sum,
            "DiscountSum": discount_sum,
            "Status": status,
        })

    try:
        import pandas as pd
        import os
        import tempfile
        from aiogram.types import FSInputFile

        report_df = pd.DataFrame(report_rows)
        data_all = await state.get_data()
        src_file = data_all.get("fuel_filename") or "fuel.xlsx"
        base = os.path.splitext(str(src_file))[0]

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = tmp.name
        tmp.close()

        report_df.to_excel(tmp_path, index=False)
        # Status ustunini bo'yash: TOPILDI=green, TOPILMADI=red
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            wb = load_workbook(tmp_path)
            ws_rep = wb.active
            status_col_idx = None
            for c in range(1, ws_rep.max_column + 1):
                if str(ws_rep.cell(row=1, column=c).value).strip() == "Status":
                    status_col_idx = c
                    break

            if status_col_idx:
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for r in range(2, ws_rep.max_row + 1):
                    cell = ws_rep.cell(row=r, column=status_col_idx)
                    v = str(cell.value).strip().upper() if cell.value is not None else ""
                    if v == "TOPILDI":
                        cell.fill = green_fill
                    elif v == "TOPILMADI":
                        cell.fill = red_fill

                wb.save(tmp_path)
        except Exception:
            pass
        await callback.message.answer_document(FSInputFile(tmp_path))
        os.remove(tmp_path)
    except Exception:
        pass

    await callback.message.edit_text("✅ Fuel yozildi. (Excel report yuborildi)")
    await state.set_state(BotStates.Fuel)


@dp.callback_query(F.data.startswith("toll_sheet:"), BotStates.TollSheetSelect)
async def callback_toll_sheet(callback: types.CallbackQuery, state: FSMContext):
    await callback.answer()
    import re
    from datetime import datetime, date

    data = await state.get_data()
    company = data.get("selected_company") or get_company(callback.from_user.id)
    if not company:
        await callback.message.edit_text("Iltimos, /start bosing va Load tanlang.")
        return
    sheet_names = data.get("toll_sheet_names") or []
    toll_entries = data.get("toll_entries") or []
    if not sheet_names or not toll_entries:
        await callback.message.edit_text("❌ Ma'lumotlar topilmadi. Qaytadan Toll fayl yuboring.")
        await state.set_state(BotStates.Toll)
        return

    idx_str = callback.data.replace("toll_sheet:", "").strip()
    try:
        idx = int(idx_str)
    except ValueError:
        return
    if idx < 0 or idx >= len(sheet_names):
        return

    sheet_name = sheet_names[idx]
    await callback.message.edit_text("⏳ Kutib turing, natija tez orada chiqadi...")

    from services.google_sheets import get_sheet_service
    sheet_service = get_sheet_service()

    ws = sheet_service.get_expenses_board(sheet_name, company)
    if not ws:
        await callback.message.edit_text(f"❌ {sheet_name} worksheet topilmadi.")
        await state.set_state(BotStates.Toll)
        return

    top_grid = ws.get("A1:Z6")
    seg_re = re.compile(r'(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})')
    year = datetime.now().year
    date_matches = []
    for r in range(min(len(top_grid), 10)):
        for c in range(min(26, len(top_grid[r]) if top_grid[r] else 0)):
            cell = top_grid[r][c] if c < len(top_grid[r]) else None
            if cell is None:
                continue
            cell_s = str(cell)
            m2 = seg_re.search(cell_s)
            if not m2:
                continue
            year_m = re.search(r'(\d{4})', cell_s)
            year_cell = int(year_m.group(1)) if year_m else year
            start_str, end_str = m2.group(1), m2.group(2)
            start_month, start_day = map(int, start_str.split('.'))
            end_month, end_day = map(int, end_str.split('.'))
            s_date = date(year_cell, start_month, start_day)
            e_year = year_cell
            if (end_month, end_day) < (start_month, start_day):
                e_year = year_cell + 1
            e_date = date(e_year, end_month, end_day)
            date_matches.append((s_date, e_date, c + 1))

    unique = {}
    for s_date, e_date, c in date_matches:
        unique[(s_date, e_date, c)] = True
    date_matches = sorted(list(unique.keys()), key=lambda x: x[2])

    transponder_cols = []
    toll_cols = []
    for r in range(min(6, len(top_grid))):
        row_data = top_grid[r] if r < len(top_grid) else []
        for c in range(min(26, len(row_data))):
            cell = row_data[c] if c < len(row_data) else None
            if cell is None:
                continue
            cell_s = str(cell).lower()
            if "transponder" in cell_s:
                transponder_cols.append(c + 1)
            if "toll" in cell_s and ("exp" in cell_s or "expenses" in cell_s):
                toll_cols.append(c + 1)

    if not toll_cols:
        toll_cols = [7]
    if not transponder_cols:
        # Eski layout fallback: D ustun
        transponder_cols = [4]

    segments = []
    if date_matches:
        for i in range(len(date_matches)):
            start_date, end_date, start_col = date_matches[i]
            end_col = (date_matches[i + 1][2] - 1) if i + 1 < len(date_matches) else 26
            t_col = next((tc for tc in toll_cols if start_col <= tc <= end_col), toll_cols[0])
            tr_col = next((tc for tc in transponder_cols if start_col <= tc <= end_col), transponder_cols[0])
            segments.append({
                "label": f"{start_date.strftime('%m.%d')}-{end_date.strftime('%m.%d')}",
                "start_date": start_date,
                "end_date": end_date,
                "transponder_col": tr_col,
                "toll_col": t_col,
            })
    else:
        segments = [{
            "label": sheet_name,
            "start_date": date(year, 1, 1),
            "end_date": date(year, 12, 31),
            "transponder_col": transponder_cols[0],
            "toll_col": toll_cols[0],
        }]

    transponder_totals_by_seg = {i: {} for i in range(len(segments))}
    matched = 0
    for item in toll_entries:
        try:
            item_date = datetime.fromisoformat(item["date"]).date()
        except Exception:
            continue
        transponder = str(item.get("transponder", "")).strip()
        if not transponder:
            continue
        toll_sum = float(item.get("toll", 0.0) or 0.0)
        for i, seg in enumerate(segments):
            if not expense_item_date_in_segment(item_date, seg["start_date"], seg["end_date"]):
                continue
            if transponder not in transponder_totals_by_seg[i]:
                transponder_totals_by_seg[i][transponder] = 0.0
            transponder_totals_by_seg[i][transponder] += toll_sum
            matched += 1
            break

    if matched == 0:
        await callback.message.edit_text("❌ Mos hafta topilmadi. Faylni tekshiring.")
        await state.set_state(BotStates.Toll)
        return

    total_updated = 0
    missing_by_seg = {}
    for i, seg in enumerate(segments):
        totals = transponder_totals_by_seg.get(i) or {}
        if not totals:
            continue
        updated, skipped, _, missing = sheet_service.update_toll_expenses(
            sheet_name,
            totals,
            toll_col=seg["toll_col"],
            transponder_col=seg.get("transponder_col", 4),
            company=company,
        )
        total_updated += updated
        missing_by_seg[i] = set(str(x) for x in (missing or []))

    # Excel report: Transponder, Week, TollSum, Status (TOPILDI=yashil, TOPILMADI=qizil)
    report_rows = []
    for i, seg in enumerate(segments):
        totals = transponder_totals_by_seg.get(i) or {}
        seg_missing = missing_by_seg.get(i) or set()
        for transponder, toll_sum in totals.items():
            trans_s = str(transponder)
            status = "TOPILMADI" if trans_s in seg_missing else "TOPILDI"
            report_rows.append({
                "Sheet": sheet_name,
                "Week": seg.get("label", ""),
                "Transponder": trans_s,
                "TollSum": toll_sum,
                "Status": status,
            })

    try:
        import pandas as pd
        import os
        import tempfile
        from aiogram.types import FSInputFile

        report_df = pd.DataFrame(report_rows)
        data_all = await state.get_data()
        src_file = data_all.get("toll_filename") or "toll.xlsx"

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = tmp.name
        tmp.close()

        report_df.to_excel(tmp_path, index=False)

        # Status ustunini bo'yash: TOPILDI=to'q yashil, TOPILMADI=to'q qizil
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill

            wb = load_workbook(tmp_path)
            ws_rep = wb.active
            status_col_idx = None
            for c in range(1, ws_rep.max_column + 1):
                if str(ws_rep.cell(row=1, column=c).value).strip() == "Status":
                    status_col_idx = c
                    break

            if status_col_idx:
                green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # to'q yashil
                red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")   # to'q qizil
                for r in range(2, ws_rep.max_row + 1):
                    cell = ws_rep.cell(row=r, column=status_col_idx)
                    v = str(cell.value).strip().upper() if cell.value is not None else ""
                    if v == "TOPILDI":
                        cell.fill = green_fill
                    elif v == "TOPILMADI":
                        cell.fill = red_fill

                wb.save(tmp_path)
        except Exception:
            pass

        await callback.message.answer_document(FSInputFile(tmp_path))
        os.remove(tmp_path)
    except Exception:
        pass

    await callback.message.edit_text("✅ Toll yozildi. Excel report yuborildi.")
    await state.set_state(BotStates.Toll)
