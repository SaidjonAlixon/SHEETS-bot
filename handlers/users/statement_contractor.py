import asyncio
import io
import os
import re
from datetime import date

import pandas as pd
from aiogram import F, types
from aiogram.fsm.context import FSMContext
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from handlers.users.statement import (
    _drivers_match,
    _extract_sheet_segments,
    _find_sheet_by_alias,
    _load_board_hint,
    _money_eq,
    _pdf_trip_ids_match_sheet_cell,
)
from keyboards.default.main_menu import get_load_select_menu
from keyboards.default.statement_menu import statement_menu
from loader import bot, dp
from services.google_sheets import get_sheet_service
from states.bot_states import BotStates
from utils.company_storage import get_company

CONTRACTOR_PROCESS_LOCK = asyncio.Lock()


def _is_gs_quota_error(err: Exception) -> bool:
    s = str(err or "")
    low = s.lower()
    return "429" in s or "quota" in low or "rate limit" in low or "resource_exhausted" in low


async def _gs_retry(fn, retries: int = 4, base_delay: float = 1.0):
    last = None
    for i in range(retries + 1):
        try:
            return fn()
        except Exception as e:
            last = e
            if not _is_gs_quota_error(e) or i >= retries:
                raise
            await asyncio.sleep(base_delay * (2**i))
    raise last


def _norm_period(s: str | None) -> str:
    text = str(s or "").strip()
    m = re.search(r"(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})", text)
    return f"{m.group(1)}-{m.group(2)}" if m else ""


def _to_amount(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s or s.lower() in ("nan", "-"):
        return None
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def _clean_name(v) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip())


def _pick_col(df: pd.DataFrame, aliases: list[str]) -> str | None:
    normalized = {str(c).strip().lower(): c for c in df.columns}
    for a in aliases:
        k = a.strip().lower()
        if k in normalized:
            return normalized[k]
    return None


def _format_result_excel(path: str) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(bold=True)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
        natija_col = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value or "").strip() == "Natija":
                natija_col = c
                break
        if natija_col:
            for r in range(2, ws.max_row + 1):
                val = str(ws.cell(row=r, column=natija_col).value or "").strip().upper()
                fill = green_fill if val == "MOS KELDI" else red_fill if val == "MOS KELMADI" else None
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = fill
        for col in range(1, ws.max_column + 1):
            max_len = 10
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=col).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 56)
    wb.save(path)


@dp.message(F.text == "👷 Contractor")
async def ask_contractor_excel(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    await state.set_state(BotStates.StatementContractorPdf)
    await message.answer(
        _load_board_hint(company)
        + "👷 <b>Contractor</b> tekshiruvi.\n\n"
        + "Iltimos, <b>Excel</b> fayl yuboring (xlsx/xls).\n"
        + "Kerakli listlar:\n"
        + "• <b>Trips Data</b>: Driver Name, Load ID, Rate (Gross), Work Period\n"
        + "• <b>P&amp;L Per Truck</b>: Driver Name, Fuel Cost, Toll Cost, Work Period",
        parse_mode="HTML",
        reply_markup=statement_menu,
    )


@dp.message(F.text == "⬅️ Back (Main Menu)", BotStates.StatementContractorPdf)
async def back_contractor_excel(message: types.Message, state: FSMContext):
    await state.set_state(BotStates.Statement)
    await message.answer("Statement Check bo'limi.", reply_markup=statement_menu)


@dp.message(F.document, BotStates.StatementContractorPdf)
async def handle_contractor_excel(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    doc = message.document
    fname = (doc.file_name or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".xls")):
        await message.answer("Iltimos, faqat Excel (xlsx/xls) yuboring.")
        return

    if CONTRACTOR_PROCESS_LOCK.locked():
        await message.answer("📥 Excel navbatga qo'shildi. Oldingi fayl tugagach ketma-ket tekshiriladi.")
    await CONTRACTOR_PROCESS_LOCK.acquire()
    try:
        await message.answer(
            f"Excel qabul qilindi. <b>{company}</b> bo'yicha Contractor tekshiruv... ⏳",
            parse_mode="HTML",
        )
        f = await bot.get_file(doc.file_id)
        content = (await bot.download_file(f.file_path)).read()
        excel = pd.read_excel(io.BytesIO(content), sheet_name=None)
        trips_sheet = next((k for k in excel.keys() if str(k).strip().lower() == "trips data"), None)
        pnl_sheet = next((k for k in excel.keys() if str(k).strip().lower() == "p&l per truck"), None)
        if not trips_sheet or not pnl_sheet:
            await message.answer("Excelda Trips Data va P&amp;L Per Truck listlari bo'lishi kerak.")
            return

        trips_df = excel[trips_sheet].copy()
        pnl_df = excel[pnl_sheet].copy()
        t_driver = _pick_col(trips_df, ["driver name", "driver"])
        t_load = _pick_col(trips_df, ["load id", "load"])
        t_rate = _pick_col(trips_df, ["rate (gross)", "rate gross", "rate"])
        t_period = _pick_col(trips_df, ["work period", "period"])
        p_driver = _pick_col(pnl_df, ["driver name", "driver"])
        p_fuel = _pick_col(pnl_df, ["fuel cost", "fuel"])
        p_toll = _pick_col(pnl_df, ["toll cost", "toll"])
        p_period = _pick_col(pnl_df, ["work period", "period"])
        if not all([t_driver, t_load, t_rate, t_period, p_driver, p_fuel, p_toll, p_period]):
            await message.answer("Kerakli ustunlar topilmadi. Rasmdagi headerlar bilan yuboring.")
            return

        sheet_service = get_sheet_service()
        load_sheet_names = await _gs_retry(lambda: sheet_service.get_all_sheet_names(company=company))
        period_to_load_sheet = {}
        for n in load_sheet_names:
            p = _norm_period(n)
            if p and p not in period_to_load_sheet:
                period_to_load_sheet[p] = n
        recent_sheets = await _gs_retry(lambda: sheet_service.get_last_n_week_sheets(14, company=company))
        index_cache = {
            sn: await _gs_retry(lambda sn=sn: sheet_service.get_load_row_index(sn, company=company))
            for sn in recent_sheets
        }

        load_rows: list[dict] = []
        fields_cache: dict[tuple[str, int], dict | None] = {}
        load_ok = load_bad = load_miss = 0
        for _, row in trips_df.iterrows():
            driver = _clean_name(row.get(t_driver))
            load_id = _clean_name(row.get(t_load))
            rate = _to_amount(row.get(t_rate))
            period_key = _norm_period(row.get(t_period))
            preferred = period_to_load_sheet.get(period_key)
            candidates = [s for s in [preferred] + recent_sheets if s]
            row_num = None
            sheet_name = None
            key = sheet_service._normalize_load_num(load_id) if load_id else ""
            for sn in candidates:
                idx = index_cache.get(sn) or {}
                if key and key in idx:
                    row_num, sheet_name = idx[key], sn
                    break

            if not row_num or not sheet_name:
                load_miss += 1
                load_rows.append(
                    {
                        "Driver": driver or "-",
                        "PDF Load ID": load_id or "-",
                        "PDF Rate": rate,
                        "Sheet": preferred or "-",
                        "Row": "-",
                        "Sheet Load ID": "-",
                        "Sheet Rate": "-",
                        "Natija": "MOS KELMADI",
                        "Sabab": "Load boardda topilmadi",
                    }
                )
                continue

            fkey = (sheet_name, int(row_num))
            if fkey not in fields_cache:
                fields_cache[fkey] = await _gs_retry(
                    lambda: sheet_service.get_settlement_compare_fields(row_num, sheet_name, company=company)
                )
            fields = fields_cache.get(fkey)
            if not fields:
                load_bad += 1
                load_rows.append(
                    {
                        "Driver": driver or "-",
                        "PDF Load ID": load_id or "-",
                        "PDF Rate": rate,
                        "Sheet": sheet_name,
                        "Row": row_num,
                        "Sheet Load ID": "-",
                        "Sheet Rate": "-",
                        "Natija": "MOS KELMADI",
                        "Sabab": "Sheet qatori o'qilmadi",
                    }
                )
                continue

            # Contractor Excelda bitta load ID yuborilishi mumkin.
            # Sheet katagida esa bir nechta ID bo'lishi mumkin (masalan 2053519//2053518).
            # Shu holatda PDF load ID sheet ichida qatnashsa - ID mos deb olinadi.
            sheet_tokens = sheet_service.split_load_cell_tokens(fields.get("load_number"))
            norm_load = sheet_service._normalize_load_num(load_id)
            id_ok = bool(norm_load and sheet_tokens and norm_load in sheet_tokens)
            if not sheet_tokens:
                id_ok = _pdf_trip_ids_match_sheet_cell(sheet_service, [load_id], fields.get("load_number"))
            rate_ok = _money_eq(rate, fields.get("rate"))
            drv_ok = _drivers_match(driver, fields.get("driver") or "") if driver else True
            fallback = id_ok and rate_ok and not drv_ok
            if id_ok and rate_ok and (drv_ok or fallback):
                load_ok += 1
                natija = "MOS KELDI"
                sabab = "Driver nomi mos emas, load id bo'yicha fallback mos keldi" if fallback else "-"
            else:
                load_bad += 1
                bits = []
                if not id_ok:
                    bits.append("Load ID mos emas")
                if not rate_ok:
                    bits.append(f"Rate mos emas (PDF {rate} vs Sheet {fields.get('rate')})")
                if not drv_ok:
                    bits.append("Driver mos emas")
                natija, sabab = "MOS KELMADI", ", ".join(bits)
            load_rows.append(
                {
                    "Driver": driver or "-",
                    "PDF Load ID": load_id or "-",
                    "PDF Rate": rate,
                    "Sheet": sheet_name,
                    "Row": row_num,
                    "Sheet Load ID": fields.get("load_number"),
                    "Sheet Rate": fields.get("rate"),
                    "Natija": natija,
                    "Sabab": sabab,
                }
            )

        exp_sheet_names = await _gs_retry(lambda: sheet_service.get_expenses_all_sheet_names(company))
        expense_candidates = [
            _find_sheet_by_alias(exp_sheet_names, "Owner Operators"),
            _find_sheet_by_alias(exp_sheet_names, "Company Drivers"),
            _find_sheet_by_alias(exp_sheet_names, "TERMINATED"),
        ]
        expense_candidates = [s for s in expense_candidates if s]
        fuel_toll_rows: list[dict] = []
        year_now = pd.Timestamp.now().year

        # Quota urilmasligi uchun expenses listlarini 1 marta o'qib cache qilamiz.
        expense_contexts: list[dict] = []
        for sh_name in expense_candidates:
            ws = await _gs_retry(lambda sh_name=sh_name: sheet_service.get_expenses_board(sh_name, company))
            if not ws:
                continue
            segs = _extract_sheet_segments(ws, year_now)
            if not segs:
                continue
            top = await _gs_retry(lambda ws=ws: ws.get("A1:Z10"))
            fuel_cols, toll_cols = [], []
            for r in range(min(10, len(top))):
                row_data = top[r] if r < len(top) else []
                for c in range(min(26, len(row_data))):
                    txt = str(row_data[c] or "").strip().lower()
                    if "fuel" in txt and ("exp" in txt or "amount" in txt or "after" in txt):
                        fuel_cols.append(c + 1)
                    if "toll" in txt and ("exp" in txt or "amount" in txt):
                        toll_cols.append(c + 1)
            if not fuel_cols:
                fuel_cols = [5]
            if not toll_cols:
                toll_cols = [7]
            names = await _gs_retry(lambda ws=ws: ws.col_values(2))
            resolved = []
            cur = ""
            for raw in names:
                nm = _clean_name(raw)
                if nm and len(nm) >= 3 and not re.match(r"^[\d\s$.,\-–—%/]+$", nm):
                    cur = nm
                resolved.append(cur)
            expense_contexts.append(
                {
                    "sheet_name": sh_name,
                    "ws": ws,
                    "segments": segs,
                    "fuel_cols": fuel_cols,
                    "toll_cols": toll_cols,
                    "resolved_names": resolved,
                }
            )

        amount_col_cache: dict[tuple[str, int], list] = {}
        for _, row in pnl_df.iterrows():
            driver = _clean_name(row.get(p_driver))
            period_key = _norm_period(row.get(p_period))
            fuel_pdf = _to_amount(row.get(p_fuel))
            toll_pdf = _to_amount(row.get(p_toll))
            for check_name, pdf_amt in (("Fuel Total", fuel_pdf), ("Toll Total", toll_pdf)):
                if pdf_amt is None:
                    continue
                best = None
                for ctx in expense_contexts:
                    ws = ctx["ws"]
                    segs = ctx["segments"]
                    seg = segs[0]
                    for sg in segs:
                        key = f"{sg[0].month:02d}.{sg[0].day:02d}-{sg[1].month:02d}.{sg[1].day:02d}"
                        if period_key and key == period_key:
                            seg = sg
                            break
                    seg_idx = segs.index(seg)
                    seg_start = seg[2]
                    seg_end = segs[seg_idx + 1][2] - 1 if seg_idx + 1 < len(segs) else 26
                    fuel_cols = ctx["fuel_cols"]
                    toll_cols = ctx["toll_cols"]
                    target_col_list = fuel_cols if check_name == "Fuel Total" else toll_cols
                    amount_col = next(
                        (cc for cc in target_col_list if seg_start <= cc <= seg_end),
                        target_col_list[0],
                    )
                    akey = (ctx["sheet_name"], int(amount_col))
                    if akey not in amount_col_cache:
                        amount_col_cache[akey] = await _gs_retry(
                            lambda ws=ws, amount_col=amount_col: ws.col_values(amount_col)
                        )
                    amounts = amount_col_cache[akey]
                    resolved = ctx["resolved_names"]
                    for i, sheet_drv in enumerate(resolved, start=1):
                        if i < 4 or not sheet_drv or not _drivers_match(driver, sheet_drv):
                            continue
                        sheet_amt = _to_amount(amounts[i - 1] if i - 1 < len(amounts) else None)
                        score = 1 if _money_eq(pdf_amt, sheet_amt) else 0
                        cand = (score, ctx["sheet_name"], i, sheet_amt)
                        if best is None or cand[0] > best[0]:
                            best = cand
                if best:
                    ok = _money_eq(pdf_amt, best[3])
                    fuel_toll_rows.append(
                        {
                            "Driver": driver or "-",
                            "Tekshiruv": check_name,
                            "Sheet": best[1],
                            "Row": best[2],
                            "PDF Amount": pdf_amt,
                            "Sheet Amount": best[3] if best[3] is not None else "-",
                            "Natija": "MOS KELDI" if ok else "MOS KELMADI",
                            "Sabab": "-" if ok else f"{check_name.lower()} mos emas",
                        }
                    )
                else:
                    fuel_toll_rows.append(
                        {
                            "Driver": driver or "-",
                            "Tekshiruv": check_name,
                            "Sheet": "-",
                            "Row": "-",
                            "PDF Amount": pdf_amt,
                            "Sheet Amount": "-",
                            "Natija": "MOS KELMADI",
                            "Sabab": "Driver topilmadi",
                        }
                    )

        load_df = pd.DataFrame(load_rows or [])
        fuel_toll_df = pd.DataFrame(fuel_toll_rows or [])
        report_name = f"Contractor_Check_Result_{message.from_user.id}_{doc.file_unique_id}.xlsx"
        with pd.ExcelWriter(report_name, engine="openpyxl") as writer:
            load_df.to_excel(writer, sheet_name="Load check", index=False)
            fuel_toll_df.to_excel(writer, sheet_name="Fuel Toll check", index=False)
        _format_result_excel(report_name)

        exp_ok = sum(1 for r in fuel_toll_rows if r.get("Natija") == "MOS KELDI")
        exp_bad = sum(1 for r in fuel_toll_rows if r.get("Natija") != "MOS KELDI")
        await message.answer(
            f"🏁 Contractor tekshiruv tugadi — <b>{company}</b>.\n"
            f"📦 Load: ✅ {load_ok} | ❌ {load_bad} | ❓ {load_miss}\n"
            f"⛽🛣️ Fuel/Toll: ✅ {exp_ok} | ❌ {exp_bad}",
            parse_mode="HTML",
        )
        await message.answer_document(types.FSInputFile(report_name))
        os.remove(report_name)
        await state.set_state(BotStates.Statement)
    except Exception as e:
        print(f"[contractor_excel_error] {e!r}")
        if _is_gs_quota_error(e):
            await message.answer("⚠️ Google Sheets limiti (429) bo'ldi. 1-2 daqiqadan keyin qayta yuboring.")
        else:
            await message.answer("Xatolik yuz berdi. Excel formatini tekshirib, qayta yuboring.")
    finally:
        await asyncio.sleep(0.8)
        CONTRACTOR_PROCESS_LOCK.release()

