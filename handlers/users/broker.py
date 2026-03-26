from aiogram import types, F
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext
from loader import dp, bot
from keyboards.default.sub_menus import broker_menu
from keyboards.default.main_menu import get_main_menu, get_load_select_menu
from gspread.exceptions import APIError as GspreadAPIError
from services.google_sheets import get_sheet_service
from services.excel_parser import ExcelParser
from states.bot_states import BotStates
from utils.company_storage import get_company
import pandas as pd
import os
import logging

logger = logging.getLogger(__name__)


@dp.message(F.text == "💰 Broker Payments")
async def enter_broker(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    await state.set_state(BotStates.Broker)
    await message.answer("Broker Payments bo'limi.\nFayl yuklang yoki buyruq tanlang:", reply_markup=broker_menu)

@dp.message(F.text == "⬅️ Orqaga", BotStates.Broker)
async def back_broker(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Asosiy menyu:", reply_markup=get_main_menu(message.from_user.id))

@dp.message(F.text == "📤 To'lov faylini yuklash", BotStates.Broker)
async def ask_broker_file(message: types.Message):
    await message.answer("Broker Payment Excel (xlsx, xls) faylini yuboring.")


# ——— 📋 Oxirgi to'lovlar ———
@dp.message(F.text == "📋 Oxirgi to'lovlar", BotStates.Broker)
async def broker_recent_payments(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    try:
        sheet_service = get_sheet_service()
        all_sheets = sheet_service.get_all_sheet_names(company)
    except Exception as e:
        if "429" in str(e):
            await message.answer("⚠️ Google Sheets limiti tugadi. 1–2 daqiqa kutib qayta urinib ko'ring.")
        else:
            await message.answer(f"Sheetga ulanish xatosi: {e}")
        return
    if not all_sheets:
        await message.answer("Listlar topilmadi.")
        return
    buttons = []
    row = []
    for i, name in enumerate(all_sheets):
        row.append(InlineKeyboardButton(text=name, callback_data=f"broker_recent:{i}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    await message.answer(
        "<b>Oxirgi to'lovlar</b>\nQaysi listdan ko'rsatay?\n\n⏳ Tugmani bosgach natija tez orada chiqadi.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )


@dp.callback_query(F.data.startswith("broker_recent:"))
async def callback_broker_recent(callback: types.CallbackQuery):
    await callback.answer()
    company = get_company(callback.from_user.id)
    if not company:
        await callback.message.edit_text("Iltimos, /start bosing va Load tanlang.")
        return
    try:
        idx = int(callback.data.replace("broker_recent:", "").strip())
    except ValueError:
        return
    await callback.message.edit_text("⏳ Kutib turing, natija tez orada chiqadi...")
    try:
        sheet_service = get_sheet_service()
        all_sheets = sheet_service.get_all_sheet_names(company)
    except Exception as e:
        await callback.message.edit_text(f"Xatolik: {e}")
        return
    if idx < 0 or idx >= len(all_sheets):
        return
    sheet_name = all_sheets[idx]
    loads = [x for x in sheet_service.get_recent_loads(sheet_name, limit=20, company=company) if x]
    # Faqat BROKER PAID to'ldirilgan qatorlarni ko'rsatish
    loads = [x for x in loads if x.get('broker_paid') and str(x.get('broker_paid')).strip() not in ('', '0', '$0.00', '-')]
    loads = loads[-15:]  # oxirgi 15 ta to'lov
    if not loads:
        await callback.message.edit_text(f"List <b>{sheet_name}</b> da oxirgi to'lovlar topilmadi.")
        return
    lines = [f"📋 <b>Oxirgi to'lovlar ({sheet_name})</b>\n"]
    for i, r in enumerate(reversed(loads), 1):
        ln = r.get('load_number') or '-'
        dr = (r.get('driver') or '-')[:18]
        paid = r.get('broker_paid') or '-'
        st = r.get('status') or '-'
        lines.append(f"{i}. LOAD #{ln} | {dr} | Paid: {paid} | {st}")
    text = "\n".join(lines)
    if len(text) > 4000:
        text = "\n".join(lines[:25]) + "\n\n... (25 ta ko'rsatildi)"
    await callback.message.edit_text(text)

@dp.message(F.document, BotStates.Broker)
async def handle_broker_document(message: types.Message):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return

    document = message.document
    file_id = document.file_id
    file_name = document.file_name

    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        await message.answer("Iltimos, faqat Excel (xlsx, xls) fayl yuklang.")
        return

    progress_msg = await message.answer("⏳ Broker fayl qabul qilindi... 10%")

    file = await bot.get_file(file_id)
    file_content = await bot.download_file(file.file_path)
    content_bytes = file_content.read()
    try:
        await progress_msg.edit_text("⏳ Broker fayl o'qilmoqda... 30%")
    except Exception:
        pass

    parsed_data = ExcelParser.parse_broker_payments_xls(content_bytes)
    if not parsed_data:
        parsed_data = ExcelParser.parse_purchase_history_report(content_bytes)

    if not parsed_data:
        await message.answer("Fayldan ma'lumot o'qib bo'lmadi. D/B (Load ID) va J (Check Amount) ustunlarini tekshiring.")
        return

    try:
        try:
            await progress_msg.edit_text("⏳ Oxirgi 10 hafta listlari tayyorlanmoqda... 50%")
        except Exception:
            pass
        sheet_service = get_sheet_service()
        sheet_names = sheet_service.get_last_n_week_sheets(n=10, company=company)
        if not sheet_names:
            await message.answer("❌ Sana oralig'i bo'lgan listlar topilmadi.")
            return
        try:
            await progress_msg.edit_text("⏳ Load # qidirilmoqda va BROKER PAID yozilmoqda... 80%")
        except Exception:
            pass
        updated, skipped, not_found, results = sheet_service.update_broker_payment_across_sheets(
            sheet_names, parsed_data, company=company
        )

        result_df = pd.DataFrame(results)
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        report_filename = tmp.name
        tmp.close()
        result_df.to_excel(report_filename, index=False)

        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            wb = load_workbook(report_filename)
            ws_rep = wb.active
            status_col = None
            for c in range(1, ws_rep.max_column + 1):
                if str(ws_rep.cell(row=1, column=c).value or "").strip() == "Status":
                    status_col = c
                    break
            if status_col:
                green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                yellow_fill = PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid")
                red_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
                for r in range(2, ws_rep.max_row + 1):
                    cell = ws_rep.cell(row=r, column=status_col)
                    v = str(cell.value or "").upper()
                    if "FOUND" in v and "NOT FOUND" not in v:
                        cell.fill = green_fill
                    elif "ALREADY FILLED" in v:
                        cell.fill = yellow_fill
                    elif "NOT FOUND" in v or "EMPTY" in v or "PROTECTED" in v:
                        cell.fill = red_fill
                wb.save(report_filename)
        except Exception:
            pass

        sheet_list = ", ".join(sheet_names[:5]) + (" ..." if len(sheet_names) > 5 else "")
        await message.answer(
            f"✅ Broker Payments yakunlandi (oxirgi 10 hafta: {sheet_list})\n\n"
            f"Yangilandi: {updated}\nO'tkazib yuborildi: {skipped}\nTopilmadi: {not_found}",
            reply_markup=broker_menu
        )
        await message.answer_document(types.FSInputFile(report_filename))
        os.remove(report_filename)
        try:
            await progress_msg.edit_text("✅ Broker Payments yakunlandi... 100%")
        except Exception:
            pass
    except GspreadAPIError as e:
        if "429" in str(e):
            await message.answer("⚠️ Google Sheets limiti tugadi. 1-2 daqiqa kutib qayta yuboring.")
        else:
            await message.answer(f"Sheet xatolik: {e}")
        return
    except Exception as e:
        logger.exception("Broker: sheet init xatosi")
        await message.answer(f"Xatolik: {e}")
        return
