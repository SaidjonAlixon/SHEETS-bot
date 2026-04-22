import os
import re
from collections import defaultdict

import pandas as pd
from aiogram import F, types
from aiogram.fsm.context import FSMContext
from openpyxl.styles import PatternFill

import config
from handlers.users.expenses import expense_item_date_in_segment, parse_toll_posting_date
from keyboards.default.main_menu import get_load_select_menu, get_main_menu
from keyboards.default.statement_menu import statement_menu
from loader import bot, dp
from services.company_driver_pdf import parse_company_driver_settlement_pdf_ai
from services.excel_parser import ExcelParser
from services.google_sheets import get_sheet_service
from states.bot_states import BotStates
from utils.company_storage import get_company


def _load_board_hint(company: str) -> str:
    """Qaysi kompaniya Load Board (Google Sheet) ishlatilayotganini ko'rsatadi."""
    return (
        f"📋 <b>Load Board:</b> {company}\n"
        f"<i>Tekshiruv shu kompaniya uchun .env dagi LOAD sheet kaliti bo'yicha.</i>\n\n"
    )


def _drivers_match(pdf_driver: str, sheet_driver: str) -> bool:
    p = re.sub(r"\s+", " ", (pdf_driver or "").strip().lower())
    s = re.sub(r"\s+", " ", (sheet_driver or "").strip().lower())
    if not p or not s:
        return False
    if p == s or p in s or s in p:
        return True
    wp, ws = set(p.split()), set(s.split())
    return len(wp & ws) >= 2


def _pdf_sheet_id_match(sheet_service, pdf_tid, sheet_load_raw) -> str:
    """PDF trip ID va sheet D ustunidagi LOAD # bir xil (normalizatsiyadan keyin)mi."""
    if not pdf_tid or sheet_load_raw is None or str(sheet_load_raw).strip() == "":
        return "-"
    a = sheet_service._normalize_load_num(pdf_tid)
    toks = sheet_service.split_load_cell_tokens(sheet_load_raw)
    if toks:
        return "ha" if a in toks else "yo'q"
    b = sheet_service._normalize_load_num(sheet_load_raw)
    if not a or not b:
        return "-"
    return "ha" if a == b else "yo'q"


def _pdf_trip_ids_match_sheet_cell(sheet_service, trip_ids: list, sheet_load_raw) -> bool:
    """Bir nechta PDF load ID sheet katakidagi // bilan birlashtirilgan ro'yxat bilan to'liq mosmi."""
    if sheet_load_raw is None or str(sheet_load_raw).strip() == "":
        return False
    sheet_toks = sheet_service.split_load_cell_tokens(sheet_load_raw)
    pdf_toks = sorted(
        sheet_service._normalize_load_num(t) for t in trip_ids if t
    )
    pdf_toks = [x for x in pdf_toks if x]
    if not pdf_toks:
        return False
    if not sheet_toks:
        one = sheet_service._normalize_load_num(sheet_load_raw)
        return bool(one) and pdf_toks == [one]
    return sorted(sheet_toks) == pdf_toks


def _find_sheet_by_alias(sheet_names: list[str], alias: str) -> str | None:
    target = str(alias or "").strip().lower()
    for n in sheet_names or []:
        if str(n).strip().lower() == target:
            return n
    return None


def _money_eq(a, b, eps=0.02) -> bool:
    try:
        return abs(float(a or 0) - float(b or 0)) <= eps
    except (TypeError, ValueError):
        return False


def _is_internal_trip_number(token: str | None) -> bool:
    """
    PDFdagi qizil Trips raqamlarini (masalan 71425, 71533) reportdan chiqarib tashlash.
    """
    if not token:
        return False
    s = re.sub(r"[^0-9]", "", str(token))
    return len(s) == 5 and s.startswith("7")


def _extract_sheet_segments(ws, year_fallback: int):
    top_grid = ws.get("A1:Z10")
    seg_re = re.compile(r"(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})")
    date_matches = []
    for r in range(min(len(top_grid), 10)):
        row_data = top_grid[r] if r < len(top_grid) else []
        for c in range(min(26, len(row_data))):
            cell = row_data[c] if c < len(row_data) else None
            if cell is None:
                continue
            cell_s = str(cell)
            m = seg_re.search(cell_s)
            if not m:
                continue
            year_m = re.search(r"(\d{4})", cell_s)
            y = int(year_m.group(1)) if year_m else year_fallback
            sm, sd = map(int, m.group(1).split("."))
            em, ed = map(int, m.group(2).split("."))
            start_d = pd.Timestamp(year=y, month=sm, day=sd).date()
            end_y = y + 1 if (em, ed) < (sm, sd) else y
            end_d = pd.Timestamp(year=end_y, month=em, day=ed).date()
            date_matches.append((start_d, end_d, c + 1))
    uniq = {(a, b, c): True for (a, b, c) in date_matches}
    return sorted(list(uniq.keys()), key=lambda x: x[2])


@dp.message(F.text == "📊 Statement Check")
async def enter_statement(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return
    await state.set_state(BotStates.Statement)
    await message.answer(
        _load_board_hint(company)
        + "Statement Check bo'limi.\nKim uchun tekshiruvni amalga oshiramiz?",
        reply_markup=statement_menu,
        parse_mode="HTML",
    )

@dp.message(F.text == "⬅️ Back (Main Menu)", BotStates.Statement)
async def back_statement(message: types.Message, state: FSMContext):
    await state.clear()
    await message.answer("Asosiy menyu:", reply_markup=get_main_menu(message.from_user.id))


@dp.message(F.text == "⬅️ Back (Main Menu)", BotStates.StatementCompanyDriverPdf)
async def back_company_driver_pdf(message: types.Message, state: FSMContext):
    await state.set_state(BotStates.Statement)
    await message.answer("Statement Check bo'limi.", reply_markup=statement_menu)


@dp.message(F.text == "⬅️ Back (Main Menu)", BotStates.StatementOwnerOperatorPdf)
async def back_owner_operator_pdf(message: types.Message, state: FSMContext):
    await state.set_state(BotStates.Statement)
    await message.answer("Statement Check bo'limi.", reply_markup=statement_menu)


def _statement_fayl_yuklash_texts():
    """Eski klaviaturalarda 📥, yangisida 📤 — ikkalasini ham qabul qilamiz."""
    return ("📤 Fayl yuklash", "📥 Fayl yuklash")


@dp.message(F.text == "🏢 Company Driver")
async def ask_company_driver_pdf(message: types.Message, state: FSMContext):
    """
    Statement FSM yo'qolganda ham ishlashi kerak (MemoryStorage, bot qayta ishga tushganda).
    Holat filterisiz — kompaniya (Load) tanlangan bo'lishi kerak.
    """
    company = get_company(message.from_user.id)
    if not company:
        await message.answer(
            "Iltimos, avval Load tanlang:",
            reply_markup=get_load_select_menu(message.from_user.id),
        )
        return
    await state.set_state(BotStates.StatementCompanyDriverPdf)
    await message.answer(
        _load_board_hint(company)
        + "🏢 <b>Company Driver</b> settlement tekshiruvi.\n\n"
        "Iltimos, <b>PDF</b> faylini yuboring.\n"
        "Fayldan haydovchi ismi, foiz (Percent), ish davri va Trips jadvalidagi "
        "load/trip ID hamda Rate (Gross) o'qiladi, keyin <b>yuqoridagi kompaniya</b> Load Board bilan solishtiriladi.",
        parse_mode="HTML",
        reply_markup=statement_menu,
    )


@dp.message(F.text == "🚚 Owner Operator")
async def ask_owner_operator_pdf(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer(
            "Iltimos, avval Load tanlang:",
            reply_markup=get_load_select_menu(message.from_user.id),
        )
        return
    await state.set_state(BotStates.StatementOwnerOperatorPdf)
    await message.answer(
        _load_board_hint(company)
        + "🚚 <b>Owner Operator</b> settlement tekshiruvi.\n\n"
        "Iltimos, <b>PDF</b> faylini yuboring.\n"
        "Tekshiruvlar:\n"
        "• Bosh qismdagi load/rate/work period → Load Board (LOAD_KEY)\n"
        "• Fuel Transaction totals → Deduction board (EXPENSES_KEY)\n"
        "• Toll Transaction (Device ID) → Deduction board (EXPENSES_KEY)\n\n"
        "Natija bitta Excelda alohida listlarda yuboriladi.",
        parse_mode="HTML",
        reply_markup=statement_menu,
    )


@dp.message(F.text.in_(_statement_fayl_yuklash_texts()), BotStates.StatementCompanyDriverPdf)
async def ask_statement_file_from_pdf_flow(message: types.Message, state: FSMContext):
    await state.set_state(BotStates.Statement)
    await message.answer("Statement Excel (xlsx, xls) faylini yuboring.")


@dp.message(F.text.in_(_statement_fayl_yuklash_texts()), BotStates.Statement)
async def ask_statement_file(message: types.Message):
    await message.answer("Statement Excel (xlsx, xls) faylini yuboring.")

@dp.message(F.document, BotStates.Statement)
async def handle_statement_doc(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer("Iltimos, avval Load tanlang:", reply_markup=get_load_select_menu(message.from_user.id))
        return

    document = message.document
    file_id = document.file_id
    file_name = document.file_name

    if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
        await message.answer("Iltimos, faqat Excel (xlsx, xls) fayl yuklang.")
        return
    
    await message.answer(
        f"Fayl qabul qilindi. <b>{company}</b> Load Board bo'yicha solishtirish boshlanmoqda... ⏳",
        parse_mode="HTML",
    )
    
    try:
        file = await bot.get_file(file_id)
        file_content = await bot.download_file(file.file_path)
        content_bytes = file_content.read()
        
        # Parse STATEMENT data
        # We reuse parse_invoice but it might need to be more flexible if columns differ.
        # Assuming Statement has 'Load #' and at least 'Total' or 'Amount'.
        
        parsed_data = ExcelParser.parse_invoice(content_bytes)
        
        if not parsed_data:
             await message.answer("Faylni o'qib bo'lmadi. 'Load #' ustuni borligiga ishonch hosil qiling.")
             return

        try:
            sheet_service = get_sheet_service()
        except Exception as e:
            err = str(e)
            if "429" in err or "Quota" in err or "quota" in err:
                await message.answer(
                    "⚠️ Google Sheets daqiqalik limiti. 1–3 daqiqa kutib qayta yuboring."
                )
            else:
                await message.answer(f"Xatolik: {e}")
            return

        results = []
        match_count = 0
        mismatch_count = 0
        not_found_count = 0
        
        sheet_cache = {}
        
        status_msg = await message.answer(
            f"📋 {company} | Jarayon: 0/{len(parsed_data)}"
        )
        
        for idx, row in enumerate(parsed_data):
            load_num = row['load_number']
            stmt_amount = row['amount']
            date_obj = row.get('date')
            
            if idx % 5 == 0:
                 try: await status_msg.edit_text(f"📋 {company} | Jarayon: {idx}/{len(parsed_data)}")
                 except: pass

            sheet_name = None
            if date_obj:
                if date_obj not in sheet_cache:
                    sheet_name = sheet_service.get_sheet_by_date(date_obj, company=company)
                    sheet_cache[date_obj] = sheet_name
                else:
                    sheet_name = sheet_cache[date_obj]
            
            # Agar sana bo'lmasa, barcha oxirgi sheetlarni qidirish qiyin. 
            # Hozircha STOP. Sana bo'lishi shart deb hisoblaymiz yoki "Date Not Found"
            
            sheet_status = "SHEET NOT FOUND"
            diff = 0
            sheet_amount = 0
            comment = ""
            
            if sheet_name:
                row_num = sheet_service.find_load_row(load_num, sheet_name, company=company)
                
                if row_num:
                    details = sheet_service.get_load_details(row_num, sheet_name, company=company)
                    if details:
                        # Compare logic
                        # Statement amount usually matches Invoiced OR Broker Paid
                        # Let's compare with Invoiced Amount (Col P) as primary
                        
                        sheet_invoiced = details['invoiced']
                        sheet_paid = details['broker_paid']
                        
                        # Qaysi biri bilan solishtirishni aniqlash qiyin bo'lishi mumkin.
                        # Odatda Statement bu bizga to'lanishi kerak bo'lgan pul (Factoring Statement)
                        # yoki Driverga to'lanadigan (Settlement).
                        # "Statement Check" deganda odatda Factoring kompaniyasi yuborgan statementni
                        # bizning Load Boarddagi Invoiced Amount bilan solishtirish tushuniladi.
                        
                        # COMPARE WITH INVOICED (Col P)
                        sheet_amount = sheet_invoiced
                        diff = sheet_amount - stmt_amount
                        
                        if abs(diff) < 0.01:
                            sheet_status = "MATCH"
                            match_count += 1
                        else:
                            sheet_status = "MISMATCH"
                            mismatch_count += 1
                            comment = f"Sheet: {sheet_amount} | Stmt: {stmt_amount}"
                    else:
                        sheet_status = "ERROR READING ROW"
                else:
                    sheet_status = "LOAD NOT FOUND"
                    not_found_count += 1
            else:
                sheet_status = "SHEET NOT FOUND (NO DATE)"
                not_found_count += 1
                
            results.append({
                "Kompaniya (Load Board)": company,
                "Load #": load_num,
                "Statement Amount": stmt_amount,
                "Sheet Amount": sheet_amount,
                "Diff": diff,
                "Status": sheet_status,
                "Date": date_obj,
                "Comment": comment
            })
            
        await status_msg.delete()
        
        # Report gen
        df = pd.DataFrame(results)
        report_name = f"Statement_Result_{file_name}"
        df.to_excel(report_name, index=False)
        
        await message.answer(
            f"🏁 Solishtirish yakunlandi (<b>{company}</b> Load Board).\n"
            f"✅ Match: {match_count}\n"
            f"❌ Mismatch: {mismatch_count}\n"
            f"❓ Not Found: {not_found_count}",
            parse_mode="HTML",
        )
                             
        await message.answer_document(types.FSInputFile(report_name))
        os.remove(report_name)
        
    except Exception as e:
        err = str(e)
        if "429" in err or "Quota" in err or "quota" in err:
            await message.answer(
                "\u26a0\ufe0f Google Sheets daqiqalik limiti. 1–3 daqiqa kutib qayta yuboring."
            )
        else:
            await message.answer(f"Xatolik: {e}")


@dp.message(F.document, BotStates.StatementCompanyDriverPdf)
async def handle_company_driver_pdf(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer(
            "Iltimos, avval Load tanlang:",
            reply_markup=get_load_select_menu(message.from_user.id),
        )
        return

    document = message.document
    file_name = (document.file_name or "").lower()
    mime = (document.mime_type or "").lower()
    if not file_name.endswith(".pdf") and mime != "application/pdf":
        await message.answer("Iltimos, faqat <b>PDF</b> fayl yuboring.", parse_mode="HTML")
        return

    await message.answer(
        f"PDF qabul qilindi. <b>{company}</b> Load Board bo'yicha o'qish va solishtirish... ⏳",
        parse_mode="HTML",
    )

    try:
        file = await bot.get_file(document.file_id)
        file_content = await bot.download_file(file.file_path)
        content_bytes = file_content.read()

        parsed = parse_company_driver_settlement_pdf_ai(content_bytes)
        warnings = parsed.get("parse_warnings") or []
        pdf_driver_name = (parsed.get("driver_name") or "").strip()

        try:
            sheet_service = get_sheet_service()
        except Exception as e:
            err = str(e)
            if "429" in err or "Quota" in err or "quota" in err:
                await message.answer(
                    "⚠️ Google Sheets daqiqalik limiti. 1–3 daqiqa kutib qayta yuboring."
                )
            else:
                await message.answer(f"Xatolik: {e}")
            return

        anchor = parsed.get("anchor_date")
        primary_sheet = None
        if anchor:
            primary_sheet = sheet_service.get_sheet_by_date(anchor, company=company)
        if not primary_sheet:
            weeks = sheet_service.get_last_n_week_sheets(10, company=company)
            primary_sheet = weeks[0] if weeks else None

        fallback_sheets = sheet_service.get_last_n_week_sheets(14, company=company)
        sheets_to_index = list(dict.fromkeys([s for s in ([primary_sheet] if primary_sheet else []) + list(fallback_sheets) if s]))
        load_index_cache: dict[str, dict] = {}
        valid_load_keys: set[str] = set()
        for sn in sheets_to_index:
            idx = sheet_service.get_load_row_index(sn, company=company)
            load_index_cache[sn] = idx
            valid_load_keys.update(idx.keys())

        results = []
        ok_n = bad_n = miss_n = 0
        skipped_trip_like_n = 0
        no_trips_in_pdf = not (parsed.get("trips") or [])

        if no_trips_in_pdf:
            find_res = None
            if pdf_driver_name:
                if primary_sheet:
                    find_res = sheet_service.find_driver_rows_on_load_sheet(
                        primary_sheet, pdf_driver_name, company=company
                    )
                if not find_res:
                    for sn in fallback_sheets:
                        if sn == primary_sheet:
                            continue
                        find_res = sheet_service.find_driver_rows_on_load_sheet(
                            sn, pdf_driver_name, company=company
                        )
                        if find_res:
                            break

            sn_no = (find_res or {}).get("sheet_name") or primary_sheet or (
                fallback_sheets[0] if fallback_sheets else "-"
            )
            first_r = (find_res or {}).get("first_row")
            matched_n = len((find_res or {}).get("matched_rows") or [])
            if first_r:
                row_txt = str(first_r)
                if matched_n > 1:
                    row_txt = f"{first_r} (blokda jami {matched_n} ta qator)"
            else:
                row_txt = "-"
            sheet_drv = (find_res or {}).get("last_resolved_name") or "-"
            izoh = (
                "Bu driver ishlamagan — PDFda trip/load ID yo'q (settlement bo'yicha)."
            )
            if find_res:
                izoh += f" Sheetda haydovchi topildi: {sheet_drv}."
            else:
                if pdf_driver_name:
                    izoh += " Sheetda shu ism bilan mos qator topilmadi."
                else:
                    izoh += " PDFdan haydovchi ismi aniqlanmadi."

            results.append(
                {
                    "PDF Load ID": "-",
                    "PDF Driver": pdf_driver_name or "-",
                    "PDF Rate (Gross)": None,
                    "Sheet List": sn_no,
                    "Sheet Row": row_txt,
                    "Sheet Driver": sheet_drv,
                    "Sheet Load ID": "-",
                    f"Sheet Rate (col {config.LOAD_BOARD_RATE_COL})": "-",
                    "ID mos": "-",
                    "Rate mos": "-",
                    "Driver mos": "ha" if find_res else "yo'q",
                    "Natija": "TRIP YO'Q",
                    "Moslik": "📄",
                    "Mos kelmagan joy": izoh,
                }
            )

        grouped: dict[tuple[str, int], list[dict]] = defaultdict(list)
        orphan_trips: list[dict] = []

        for trip in parsed.get("trips") or []:
            tid = trip.get("trip_id")
            if _is_internal_trip_number(tid):
                skipped_trip_like_n += 1
                continue
            key = sheet_service._normalize_load_num(tid) if tid else ""

            row_num = None
            sn = None
            if primary_sheet and key:
                idx = load_index_cache.get(primary_sheet) or {}
                if key in idx:
                    row_num, sn = idx[key], primary_sheet
            if not row_num and key:
                for fb in fallback_sheets:
                    idx = load_index_cache.get(fb) or {}
                    if key in idx:
                        row_num, sn = idx[key], fb
                        break

            if sn and row_num:
                grouped[(sn, row_num)].append(trip)
            else:
                orphan_trips.append(trip)

        def _append_one_result(
            tid_display,
            pdf_rate_display,
            sum_pdf_rates,
            sn,
            row_num,
            sheet_load,
            sheet_rate,
            sheet_driver,
            id_mos_bool,
            rate_ok_bool,
            driver_ok_bool,
            natija,
            sabab,
        ):
            results.append(
                {
                    "PDF Load ID": tid_display,
                    "PDF Driver": pdf_driver_name or "-",
                    "PDF Rate (Gross)": pdf_rate_display,
                    "PDF Rate (jami)": sum_pdf_rates,
                    "Sheet List": sn or "(topilmadi)",
                    "Sheet Row": row_num or "-",
                    "Sheet Driver": sheet_driver if row_num else "-",
                    "Sheet Load ID": sheet_load if row_num else "-",
                    f"Sheet Rate (col {config.LOAD_BOARD_RATE_COL})": sheet_rate if row_num else "-",
                    "ID mos": "ha" if id_mos_bool else "yo'q",
                    "Rate mos": "ha" if rate_ok_bool else "yo'q",
                    "Driver mos": "ha" if driver_ok_bool else "yo'q",
                    "Natija": natija,
                    "Moslik": "✅" if natija == "MOS KELDI" else "❌",
                    "Mos kelmagan joy": sabab or "-",
                }
            )

        for (sn, row_num), trip_list in grouped.items():
            tids = [t.get("trip_id") for t in trip_list if t.get("trip_id")]
            rates = []
            for t in trip_list:
                try:
                    rates.append(float(t.get("rate_gross") or 0))
                except (TypeError, ValueError):
                    rates.append(0.0)
            sum_pdf = sum(rates)
            parts = []
            for t in trip_list:
                tid = t.get("trip_id") or "-"
                r = t.get("rate_gross")
                parts.append(f"{tid} ({r})")
            tid_display = " + ".join(parts) if len(trip_list) > 1 else (tids[0] if tids else "-")
            pdf_rate_display = (
                " + ".join(str(t.get("rate_gross")) for t in trip_list)
                if len(trip_list) > 1
                else trip_list[0].get("rate_gross")
            )

            fields = sheet_service.get_settlement_compare_fields(row_num, sn, company=company)
            if not fields:
                bad_n += 1
                _append_one_result(
                    tid_display,
                    pdf_rate_display,
                    sum_pdf,
                    sn,
                    row_num,
                    "",
                    None,
                    "",
                    False,
                    False,
                    False,
                    "MOS KELMADI",
                    "Sheet qatori o'qilmadi",
                )
                continue

            sheet_load = fields.get("load_number")
            sheet_rate = fields.get("rate")
            sheet_driver = (fields.get("driver") or "").strip()
            try:
                rate_ok = abs(float(sheet_rate or 0) - float(sum_pdf)) < 0.02
            except (TypeError, ValueError):
                rate_ok = False

            id_mos = _pdf_trip_ids_match_sheet_cell(sheet_service, tids, sheet_load)
            if pdf_driver_name:
                driver_ok = _drivers_match(pdf_driver_name, sheet_driver)
            else:
                driver_ok = True

            if id_mos and rate_ok and driver_ok:
                natija = "MOS KELDI"
                ok_n += 1
                sabab = ""
            else:
                natija = "MOS KELMADI"
                bad_n += 1
                bits = []
                if not id_mos:
                    bits.append("LOAD ID mos emas")
                if not rate_ok:
                    bits.append(
                        f"RATE mos emas (PDF jami {sum_pdf} vs Sheet {sheet_rate})"
                    )
                if not driver_ok:
                    bits.append("Driver mos emas")
                sabab = ", ".join(bits)

            _append_one_result(
                tid_display,
                pdf_rate_display,
                sum_pdf,
                sn,
                row_num,
                sheet_load,
                sheet_rate,
                sheet_driver,
                id_mos,
                rate_ok,
                driver_ok,
                natija,
                sabab,
            )

        for trip in orphan_trips:
            tid = trip.get("trip_id")
            pdf_rate = trip.get("rate_gross")
            natija = "MOS KELMADI"
            miss_n += 1
            sabab = "LOAD ID sheetda topilmadi"
            try:
                sum_or = float(pdf_rate or 0)
            except (TypeError, ValueError):
                sum_or = 0.0
            _append_one_result(
                tid or "-",
                pdf_rate,
                sum_or,
                None,
                None,
                "",
                None,
                "",
                False,
                False,
                False,
                natija,
                sabab,
            )

        detail = pd.DataFrame(results)
        base = re.sub(r'[<>:"/\\|?*]', "_", document.file_name or "report").strip() or "report"
        report_name = f"CompanyDriver_check_{base}.xlsx"
        if not report_name.lower().endswith(".xlsx"):
            report_name += ".xlsx"

        with pd.ExcelWriter(report_name, engine="openpyxl") as writer:
            detail.to_excel(writer, sheet_name="Solishtirish", index=False)
            ws = writer.book["Solishtirish"]
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            natija_col = None
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(row=1, column=c).value or "").strip() == "Natija":
                    natija_col = c
                    break
            if natija_col:
                for r in range(2, ws.max_row + 1):
                    natija_val = str(ws.cell(row=r, column=natija_col).value or "").strip().upper()
                    if natija_val == "MOS KELDI":
                        for c in range(1, ws.max_column + 1):
                            ws.cell(row=r, column=c).fill = green_fill

        warn_text = "\n".join(f"⚠️ {w}" for w in warnings) if warnings else ""
        if warn_text:
            await message.answer(warn_text)

        no_trip_line = ""
        if no_trips_in_pdf:
            no_trip_line = (
                f"\n📄 <b>PDFda trip/load yo'q.</b> Excelda haydovchi joyi va "
                f"<i>ishlamagan</i> izohi qatorida.\n"
            )

        await message.answer(
            f"🏁 Tekshiruv tugadi — <b>{company}</b> Load Board.\n"
            f"✅ To'g'ri: {ok_n}\n"
            f"❌ Noto'g'ri: {bad_n}\n"
            f"❓ Topilmadi: {miss_n}\n"
            f"🚫 Trip-no sifatida chiqarib tashlandi: {skipped_trip_like_n}\n"
            f"📋 Asosiy list: {primary_sheet or '-'}"
            f"{no_trip_line}",
            parse_mode="HTML",
        )
        await message.answer_document(types.FSInputFile(report_name))
        os.remove(report_name)
        await state.set_state(BotStates.Statement)

    except Exception as e:
        err = str(e)
        if "429" in err or "Quota" in err or "quota" in err:
            await message.answer(
                "\u26a0\ufe0f Google Sheets daqiqalik limiti. 1–3 daqiqa kutib qayta yuboring."
            )
        else:
            await message.answer(f"Xatolik: {e}")


@dp.message(F.document, BotStates.StatementOwnerOperatorPdf)
async def handle_owner_operator_pdf(message: types.Message, state: FSMContext):
    company = get_company(message.from_user.id)
    if not company:
        await message.answer(
            "Iltimos, avval Load tanlang:",
            reply_markup=get_load_select_menu(message.from_user.id),
        )
        return

    document = message.document
    file_name = (document.file_name or "").lower()
    mime = (document.mime_type or "").lower()
    if not file_name.endswith(".pdf") and mime != "application/pdf":
        await message.answer("Iltimos, faqat <b>PDF</b> fayl yuboring.", parse_mode="HTML")
        return

    await message.answer(
        f"PDF qabul qilindi. <b>{company}</b> bo'yicha Load/Fuel/Toll tekshiruv... ⏳",
        parse_mode="HTML",
    )
    try:
        file = await bot.get_file(document.file_id)
        file_content = await bot.download_file(file.file_path)
        content_bytes = file_content.read()
        parsed = parse_company_driver_settlement_pdf_ai(content_bytes)
        warnings = parsed.get("parse_warnings") or []

        pdf_driver_name = (parsed.get("driver_name") or "").strip()
        anchor = parsed.get("anchor_date")

        sheet_service = get_sheet_service()
        primary_sheet = sheet_service.get_sheet_by_date(anchor, company=company) if anchor else None
        fallback_sheets = sheet_service.get_last_n_week_sheets(14, company=company)
        sheets_to_index = list(
            dict.fromkeys(
                [s for s in ([primary_sheet] if primary_sheet else []) + list(fallback_sheets) if s]
            )
        )

        load_index_cache: dict[str, dict] = {}
        valid_load_keys: set[str] = set()
        for sn in sheets_to_index:
            idx = sheet_service.get_load_row_index(sn, company=company)
            load_index_cache[sn] = idx
            valid_load_keys.update(idx.keys())

        # ---- LOAD CHECK (owner operator) ----
        load_rows = []
        load_ok = load_bad = load_miss = 0
        grouped: dict[tuple[str, int], list[dict]] = defaultdict(list)
        orphans: list[dict] = []
        for trip in parsed.get("trips") or []:
            tid = trip.get("trip_id")
            if _is_internal_trip_number(tid):
                continue
            key = sheet_service._normalize_load_num(tid) if tid else ""
            row_num = None
            sn = None
            if primary_sheet and key:
                idx = load_index_cache.get(primary_sheet) or {}
                if key in idx:
                    row_num, sn = idx[key], primary_sheet
            if not row_num and key:
                for fb in fallback_sheets:
                    idx = load_index_cache.get(fb) or {}
                    if key in idx:
                        row_num, sn = idx[key], fb
                        break
            if row_num and sn:
                grouped[(sn, row_num)].append(trip)
            else:
                orphans.append(trip)

        for (sn, row_num), trip_list in grouped.items():
            tids = [t.get("trip_id") for t in trip_list if t.get("trip_id")]
            sum_pdf = 0.0
            for t in trip_list:
                try:
                    sum_pdf += float(t.get("rate_gross") or 0)
                except (TypeError, ValueError):
                    pass
            fields = sheet_service.get_settlement_compare_fields(row_num, sn, company=company)
            if not fields:
                load_bad += 1
                load_rows.append(
                    {
                        "Tekshiruv": "Load",
                        "Sheet": sn,
                        "Row": row_num,
                        "PDF Driver": pdf_driver_name or "-",
                        "PDF Load IDs": " + ".join(tids) if tids else "-",
                        "PDF Rate jami": sum_pdf,
                        "Sheet Load ID": "-",
                        "Sheet Rate": "-",
                        "Natija": "MOS KELMADI",
                        "Sabab": "Sheet qatori o'qilmadi",
                    }
                )
                continue
            sheet_load = fields.get("load_number")
            sheet_rate = fields.get("rate")
            sheet_driver = (fields.get("driver") or "").strip()
            id_ok = _pdf_trip_ids_match_sheet_cell(sheet_service, tids, sheet_load)
            rate_ok = _money_eq(sum_pdf, sheet_rate)
            drv_ok = _drivers_match(pdf_driver_name, sheet_driver) if pdf_driver_name else True
            if id_ok and rate_ok and drv_ok:
                load_ok += 1
                natija = "MOS KELDI"
                sabab = "-"
            else:
                load_bad += 1
                natija = "MOS KELMADI"
                bits = []
                if not id_ok:
                    bits.append("Load ID mos emas")
                if not rate_ok:
                    bits.append(f"Rate mos emas (PDF {sum_pdf} vs Sheet {sheet_rate})")
                if not drv_ok:
                    bits.append("Driver mos emas")
                sabab = ", ".join(bits)
            load_rows.append(
                {
                    "Tekshiruv": "Load",
                    "Sheet": sn,
                    "Row": row_num,
                    "PDF Driver": pdf_driver_name or "-",
                    "PDF Load IDs": " + ".join(tids) if tids else "-",
                    "PDF Rate jami": sum_pdf,
                    "Sheet Load ID": sheet_load,
                    "Sheet Rate": sheet_rate,
                    "Natija": natija,
                    "Sabab": sabab,
                }
            )
        for trip in orphans:
            load_miss += 1
            load_rows.append(
                {
                    "Tekshiruv": "Load",
                    "Sheet": "-",
                    "Row": "-",
                    "PDF Driver": pdf_driver_name or "-",
                    "PDF Load IDs": trip.get("trip_id") or "-",
                    "PDF Rate jami": trip.get("rate_gross"),
                    "Sheet Load ID": "-",
                    "Sheet Rate": "-",
                    "Natija": "MOS KELMADI",
                    "Sabab": "Load boardda topilmadi",
                }
            )

        # ---- EXPENSES CHECK (Fuel + Toll) ----
        expenses_rows = []
        exp_sheet_names = sheet_service.get_expenses_all_sheet_names(company)
        candidate_exp_sheets = [
            _find_sheet_by_alias(exp_sheet_names, "Owner Operators"),
            _find_sheet_by_alias(exp_sheet_names, "Company Drivers"),
            _find_sheet_by_alias(exp_sheet_names, "TERMINATED"),
        ]
        candidate_exp_sheets = [x for x in candidate_exp_sheets if x]
        year_now = pd.Timestamp.now().year

        fuel_total_pdf = parsed.get("fuel_total_pay_amount")
        toll_entries = parsed.get("toll_transactions") or []
        toll_pdf_by_device: dict[str, float] = defaultdict(float)
        for t in toll_entries:
            did = str(t.get("device_id") or "").strip()
            if not did:
                continue
            try:
                toll_pdf_by_device[did] += float(t.get("pay_amount") or 0)
            except (TypeError, ValueError):
                continue

        fuel_done = False
        for sh_name in candidate_exp_sheets:
            ws = sheet_service.get_expenses_board(sh_name, company)
            if not ws:
                continue
            segments = _extract_sheet_segments(ws, year_now)
            if not segments:
                continue
            top_grid = ws.get("A1:Z10")
            card_cols, fuel_cols, trans_cols, toll_cols = [], [], [], []
            for r in range(min(10, len(top_grid))):
                row_data = top_grid[r] if r < len(top_grid) else []
                for c in range(min(26, len(row_data))):
                    txt = str(row_data[c] or "").strip().lower()
                    if not txt:
                        continue
                    if ("efs" in txt and "card" in txt) or txt == "card #":
                        card_cols.append(c + 1)
                    if "fuel" in txt and ("exp" in txt or "after" in txt or "amount" in txt):
                        fuel_cols.append(c + 1)
                    if "transponder" in txt or "pptag" in txt:
                        trans_cols.append(c + 1)
                    if "toll" in txt and ("exp" in txt or "amount" in txt):
                        toll_cols.append(c + 1)
            if not fuel_cols:
                fuel_cols = [5]
            if not card_cols:
                card_cols = [3, 4]
            if not trans_cols:
                trans_cols = [4, 3]
            if not toll_cols:
                toll_cols = [7]

            name_vals = ws.col_values(2)
            resolved_names = []
            cur = ""
            for raw in name_vals:
                s = str(raw or "").strip()
                if s and len(s) >= 3 and not re.match(r"^[\d\s$.,\-–—%/]+$", s):
                    cur = s
                resolved_names.append(cur)
            matched_rows = []
            if pdf_driver_name:
                for i, drv in enumerate(resolved_names, start=1):
                    if i < 4:
                        continue
                    if drv and _drivers_match(pdf_driver_name, drv):
                        matched_rows.append(i)

            seg_match = None
            wp_start = parsed.get("work_period_start")
            for seg in segments:
                s_date, e_date, start_col = seg
                if wp_start and expense_item_date_in_segment(wp_start, s_date, e_date):
                    seg_match = seg
                    break
            if not seg_match:
                seg_match = segments[0]
            seg_idx = segments.index(seg_match)
            seg_start_col = seg_match[2]
            seg_end_col = segments[seg_idx + 1][2] - 1 if seg_idx + 1 < len(segments) else 26
            fuel_col = next((fc for fc in fuel_cols if seg_start_col <= fc <= seg_end_col), fuel_cols[0])
            trans_col = next((tc for tc in trans_cols if seg_start_col <= tc <= seg_end_col), trans_cols[0])
            toll_col = next((tc for tc in toll_cols if seg_start_col <= tc <= seg_end_col), toll_cols[0])

            # Fuel check by driver name row + week segment
            if not fuel_done and fuel_total_pdf is not None:
                if matched_rows:
                    for rr in matched_rows:
                        sheet_fuel_val = ws.cell(rr, fuel_col).value
                        sheet_fuel_num = float(str(sheet_fuel_val or "0").replace("$", "").replace(",", "") or 0)
                        ok = _money_eq(fuel_total_pdf, sheet_fuel_num)
                        expenses_rows.append(
                            {
                                "Tekshiruv": "Fuel Total",
                                "Sheet": sh_name,
                                "Row": rr,
                                "PDF Driver": pdf_driver_name or "-",
                                "Device/Identifier": "-",
                                "PDF Amount": fuel_total_pdf,
                                "Sheet Amount": sheet_fuel_num,
                                "Natija": "MOS KELDI" if ok else "MOS KELMADI",
                                "Sabab": "-" if ok else "Fuel total mos emas",
                            }
                        )
                        fuel_done = True
                        break
                elif pdf_driver_name:
                    expenses_rows.append(
                        {
                            "Tekshiruv": "Fuel Total",
                            "Sheet": sh_name,
                            "Row": "-",
                            "PDF Driver": pdf_driver_name,
                            "Device/Identifier": "-",
                            "PDF Amount": fuel_total_pdf,
                            "Sheet Amount": "-",
                            "Natija": "MOS KELMADI",
                            "Sabab": "Driver topilmadi",
                        }
                    )
                    fuel_done = True

            # Toll check by device ID/transponder in each list
            if toll_pdf_by_device:
                trans_vals = ws.col_values(trans_col)
                toll_vals = ws.col_values(toll_col)
                norm_to_row = {}
                for i, raw in enumerate(trans_vals, start=1):
                    if i < 4:
                        continue
                    nv = sheet_service._normalize_load_num(raw)
                    if nv and nv not in norm_to_row:
                        norm_to_row[nv] = i
                for device_id, pdf_amt in toll_pdf_by_device.items():
                    target = sheet_service._normalize_load_num(device_id)
                    rr = norm_to_row.get(target)
                    if rr:
                        raw_sheet_toll = toll_vals[rr - 1] if rr - 1 < len(toll_vals) else ""
                        try:
                            sheet_amt = float(str(raw_sheet_toll or "0").replace("$", "").replace(",", "") or 0)
                        except ValueError:
                            sheet_amt = 0.0
                        ok = _money_eq(pdf_amt, sheet_amt)
                        expenses_rows.append(
                            {
                                "Tekshiruv": "Toll Device",
                                "Sheet": sh_name,
                                "Row": rr,
                                "PDF Driver": pdf_driver_name or "-",
                                "Device/Identifier": device_id,
                                "PDF Amount": pdf_amt,
                                "Sheet Amount": sheet_amt,
                                "Natija": "MOS KELDI" if ok else "MOS KELMADI",
                                "Sabab": "-" if ok else "Toll amount mos emas",
                            }
                        )

        if fuel_total_pdf is None:
            expenses_rows.append(
                {
                    "Tekshiruv": "Fuel Total",
                    "Sheet": "-",
                    "Row": "-",
                    "PDF Driver": pdf_driver_name or "-",
                    "Device/Identifier": "-",
                    "PDF Amount": "-",
                    "Sheet Amount": "-",
                    "Natija": "MOS KELMADI",
                    "Sabab": "PDFda Fuel totals topilmadi",
                }
            )
        if toll_pdf_by_device and not any(r.get("Tekshiruv") == "Toll Device" for r in expenses_rows):
            for did, amt in toll_pdf_by_device.items():
                expenses_rows.append(
                    {
                        "Tekshiruv": "Toll Device",
                        "Sheet": "-",
                        "Row": "-",
                        "PDF Driver": pdf_driver_name or "-",
                        "Device/Identifier": did,
                        "PDF Amount": amt,
                        "Sheet Amount": "-",
                        "Natija": "MOS KELMADI",
                        "Sabab": "Device ID topilmadi",
                    }
                )

        # ---- EXPORT ----
        load_df = pd.DataFrame(load_rows or [])
        exp_df = pd.DataFrame(expenses_rows or [])
        base = re.sub(r'[<>:"/\\|?*]', "_", document.file_name or "report").strip() or "report"
        report_name = f"OwnerOperator_check_{base}.xlsx"
        if not report_name.lower().endswith(".xlsx"):
            report_name += ".xlsx"
        with pd.ExcelWriter(report_name, engine="openpyxl") as writer:
            load_df.to_excel(writer, sheet_name="Load check", index=False)
            exp_df.to_excel(writer, sheet_name="Fuel Toll check", index=False)
            for ws_name in ("Load check", "Fuel Toll check"):
                ws = writer.book[ws_name]
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                natija_col = None
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(row=1, column=c).value or "").strip() == "Natija":
                        natija_col = c
                        break
                if natija_col:
                    for r in range(2, ws.max_row + 1):
                        natija_val = str(ws.cell(row=r, column=natija_col).value or "").strip().upper()
                        if natija_val == "MOS KELDI":
                            for c in range(1, ws.max_column + 1):
                                ws.cell(row=r, column=c).fill = green_fill

        warn_text = "\n".join(f"⚠️ {w}" for w in warnings) if warnings else ""
        if warn_text:
            await message.answer(warn_text)
        exp_ok = sum(1 for r in expenses_rows if r.get("Natija") == "MOS KELDI")
        exp_bad = sum(1 for r in expenses_rows if r.get("Natija") != "MOS KELDI")
        await message.answer(
            f"🏁 Owner Operator tekshiruv tugadi — <b>{company}</b>.\n"
            f"📦 Load: ✅ {load_ok} | ❌ {load_bad} | ❓ {load_miss}\n"
            f"⛽🛣️ Fuel/Toll: ✅ {exp_ok} | ❌ {exp_bad}",
            parse_mode="HTML",
        )
        await message.answer_document(types.FSInputFile(report_name))
        os.remove(report_name)
        await state.set_state(BotStates.Statement)
    except Exception as e:
        err = str(e)
        if "429" in err or "Quota" in err or "quota" in err:
            await message.answer("⚠️ Google Sheets daqiqalik limiti. 1–3 daqiqa kutib qayta yuboring.")
        else:
            await message.answer(f"Xatolik: {e}")
